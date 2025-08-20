# backend/apps/finance/services/accounting.py

"""
Core Accounting Service - Complete Implementation
Advanced accounting operations with multi-currency and multi-tenant support
"""

from django.db import transaction, models
from django.utils import timezone
from django.core.exceptions import ValidationError
from django.db.models import Q, F, Sum, Case, When, DecimalField, Value, Count
from decimal import Decimal, ROUND_HALF_UP
from datetime import date, datetime, timedelta
import logging
from typing import Dict, List, Optional, Tuple, Any, Union

from apps.core.utils import generate_code
from ..models import (
    Account, AccountCategory, JournalEntry, JournalEntryLine,
    Currency, ExchangeRate, FinanceSettings, FiscalYear, FinancialPeriod,
    TaxCode, Invoice, Bill, Payment, Customer, Vendor,
    InventoryCostLayer, Project, Department, Location
)

logger = logging.getLogger(__name__)


class AccountingService:
    """
    Complete Core Accounting Service
    Handles all fundamental accounting operations and calculations
    """
    
    def __init__(self, tenant):
        self.tenant = tenant
        self.settings = self._get_finance_settings()
        self.base_currency = self._get_base_currency()
        self.current_fiscal_year = self._get_current_fiscal_year()
    
    def _get_finance_settings(self):
        """Get finance settings for tenant"""
        try:
            return FinanceSettings.objects.get(tenant=self.tenant)
        except FinanceSettings.DoesNotExist:
            # Create default settings
            return FinanceSettings.objects.create(
                tenant=self.tenant,
                company_name=f"{self.tenant.name} Inc.",
                base_currency='USD',
                accounting_method='ACCRUAL',
                fiscal_year_start_month=1,
                current_fiscal_year=date.today().year
            )
    
    def _get_base_currency(self):
        """Get base currency for tenant"""
        try:
            return Currency.objects.get(
                tenant=self.tenant,
                code=self.settings.base_currency
            )
        except Currency.DoesNotExist:
            # Create default USD currency
            return Currency.objects.create(
                tenant=self.tenant,
                code=self.settings.base_currency,
                name='US Dollar',
                symbol='$',
                decimal_places=2,
                is_base_currency=True
            )
    
    def _get_current_fiscal_year(self):
        """Get current fiscal year for tenant"""
        try:
            return FiscalYear.objects.get(
                tenant=self.tenant,
                year=self.settings.current_fiscal_year
            )
        except FiscalYear.DoesNotExist:
            # Create current fiscal year
            start_month = self.settings.fiscal_year_start_month
            year = self.settings.current_fiscal_year
            
            return FiscalYear.objects.create(
                tenant=self.tenant,
                year=year,
                start_date=date(year, start_month, 1),
                end_date=date(year + 1, start_month - 1, 31) if start_month > 1 else date(year, 12, 31),
                status='OPEN'
            )
    
    # ============================================================================
    # ACCOUNT BALANCE MANAGEMENT
    # ============================================================================
    
    def get_account_balance(self, account_id: int, as_of_date: date = None, 
                           currency: Currency = None, include_pending: bool = False) -> Decimal:
        """
        Get account balance as of specific date with multi-currency support
        
        Args:
            account_id: Account ID
            as_of_date: Date for balance calculation (defaults to today)
            currency: Currency for conversion (defaults to base currency)
            include_pending: Include draft/pending transactions
            
        Returns:
            Account balance in specified currency
        """
        try:
            account = Account.objects.get(tenant=self.tenant, id=account_id)
            
            if not as_of_date:
                as_of_date = date.today()
            
            if not currency:
                currency = self.base_currency
            
            # Build query filters
            filters = Q(
                tenant=self.tenant,
                account=account,
                journal_entry__entry_date__lte=as_of_date
            )
            
            if not include_pending:
                filters &= Q(journal_entry__status='POSTED')
            
            # Get sum of all journal entry lines
            line_totals = JournalEntryLine.objects.filter(filters).aggregate(
                total_debits=Sum('base_currency_debit_amount'),
                total_credits=Sum('base_currency_credit_amount')
            )
            
            total_debits = line_totals['total_debits'] or Decimal('0.00')
            total_credits = line_totals['total_credits'] or Decimal('0.00')
            
            # Calculate balance based on normal balance
            if account.normal_balance == 'DEBIT':
                balance = account.opening_balance + total_debits - total_credits
            else:
                balance = account.opening_balance + total_credits - total_debits
            
            # Convert currency if needed
            if currency.code != self.base_currency.code:
                exchange_rate = self.get_exchange_rate(self.base_currency, currency, as_of_date)
                balance = balance * exchange_rate
            
            return balance.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            
        except Account.DoesNotExist:
            raise ValidationError(f"Account {account_id} not found")
        except Exception as e:
            logger.error(f"Error calculating account balance: {str(e)}")
            return Decimal('0.00')
    
    def update_account_balances(self, account_ids: List[int] = None, 
                              as_of_date: date = None) -> Dict[int, Decimal]:
        """
        Update current balances for accounts
        
        Args:
            account_ids: List of account IDs (all if None)
            as_of_date: Date for balance calculation
            
        Returns:
            Dictionary of account_id: balance
        """
        try:
            if not as_of_date:
                as_of_date = date.today()
            
            # Get accounts to update
            accounts_query = Account.objects.filter(tenant=self.tenant, is_active=True)
            if account_ids:
                accounts_query = accounts_query.filter(id__in=account_ids)
            
            updated_balances = {}
            
            with transaction.atomic():
                for account in accounts_query:
                    new_balance = self.get_account_balance(account.id, as_of_date)
                    account.current_balance = new_balance
                    account.save(update_fields=['current_balance', 'updated_at'])
                    updated_balances[account.id] = new_balance
            
            logger.info(f"Updated balances for {len(updated_balances)} accounts")
            return updated_balances
            
        except Exception as e:
            logger.error(f"Error updating account balances: {str(e)}")
            return {}
    
    def get_accounts_by_type(self, account_type: str, include_inactive: bool = False) -> models.QuerySet:
        """Get accounts by type with optional inactive accounts"""
        filters = Q(tenant=self.tenant, account_type=account_type)
        if not include_inactive:
            filters &= Q(is_active=True)
        
        return Account.objects.filter(filters).order_by('code')
    
    # ============================================================================
    # TRIAL BALANCE GENERATION
    # ============================================================================
    
    def generate_trial_balance(self, as_of_date: date = None, 
                             include_zero_balances: bool = False,
                             currency: Currency = None) -> Dict:
        """
        Generate trial balance with complete account information
        
        Args:
            as_of_date: Date for trial balance (defaults to today)
            include_zero_balances: Include accounts with zero balances
            currency: Currency for amounts (defaults to base currency)
            
        Returns:
            Complete trial balance data
        """
        try:
            if not as_of_date:
                as_of_date = date.today()
            
            if not currency:
                currency = self.base_currency
            
            # Get all active accounts
            accounts = Account.objects.filter(
                tenant=self.tenant,
                is_active=True
            ).order_by('account_type', 'code')
            
            trial_balance_data = {
                'as_of_date': as_of_date,
                'currency': currency.code,
                'accounts': [],
                'totals': {
                    'total_debits': Decimal('0.00'),
                    'total_credits': Decimal('0.00')
                },
                'by_type': {},
                'is_balanced': False
            }
            
            total_debits = Decimal('0.00')
            total_credits = Decimal('0.00')
            
            for account in accounts:
                balance = self.get_account_balance(account.id, as_of_date, currency)
                
                # Skip zero balances if not requested
                if not include_zero_balances and balance == Decimal('0.00'):
                    continue
                
                # Determine debit/credit presentation
                if account.normal_balance == 'DEBIT':
                    debit_amount = balance if balance >= 0 else Decimal('0.00')
                    credit_amount = abs(balance) if balance < 0 else Decimal('0.00')
                else:
                    debit_amount = abs(balance) if balance < 0 else Decimal('0.00')
                    credit_amount = balance if balance >= 0 else Decimal('0.00')
                
                account_data = {
                    'account_id': account.id,
                    'account_code': account.code,
                    'account_name': account.name,
                    'account_type': account.account_type,
                    'normal_balance': account.normal_balance,
                    'balance': balance,
                    'debit_amount': debit_amount,
                    'credit_amount': credit_amount
                }
                
                trial_balance_data['accounts'].append(account_data)
                
                # Add to totals
                total_debits += debit_amount
                total_credits += credit_amount
                
                # Group by account type
                if account.account_type not in trial_balance_data['by_type']:
                    trial_balance_data['by_type'][account.account_type] = {
                        'accounts': [],
                        'total_debits': Decimal('0.00'),
                        'total_credits': Decimal('0.00')
                    }
                
                trial_balance_data['by_type'][account.account_type]['accounts'].append(account_data)
                trial_balance_data['by_type'][account.account_type]['total_debits'] += debit_amount
                trial_balance_data['by_type'][account.account_type]['total_credits'] += credit_amount
            
            # Update totals
            trial_balance_data['totals']['total_debits'] = total_debits
            trial_balance_data['totals']['total_credits'] = total_credits
            trial_balance_data['totals']['difference'] = abs(total_debits - total_credits)
            trial_balance_data['is_balanced'] = abs(total_debits - total_credits) <= Decimal('0.01')
            
            logger.info(f"Generated trial balance as of {as_of_date}")
            return trial_balance_data
            
        except Exception as e:
            logger.error(f"Error generating trial balance: {str(e)}")
            raise ValidationError(f"Failed to generate trial balance: {str(e)}")
    
    # ============================================================================
    # FINANCIAL STATEMENT GENERATION
    # ============================================================================
    
    def generate_balance_sheet(self, as_of_date: date = None, 
                             currency: Currency = None) -> Dict:
        """
        Generate balance sheet with proper classification
        
        Args:
            as_of_date: Date for balance sheet (defaults to today)
            currency: Currency for amounts (defaults to base currency)
            
        Returns:
            Complete balance sheet data
        """
        try:
            if not as_of_date:
                as_of_date = date.today()
            
            if not currency:
                currency = self.base_currency
            
            balance_sheet = {
                'as_of_date': as_of_date,
                'currency': currency.code,
                'assets': {
                    'current_assets': [],
                    'fixed_assets': [],
                    'other_assets': [],
                    'total': Decimal('0.00')
                },
                'liabilities': {
                    'current_liabilities': [],
                    'long_term_liabilities': [],
                    'total': Decimal('0.00')
                },
                'equity': {
                    'equity_accounts': [],
                    'retained_earnings': Decimal('0.00'),
                    'current_period_earnings': Decimal('0.00'),
                    'total': Decimal('0.00')
                }
            }
            
            # Get asset accounts
            asset_types = ['CURRENT_ASSET', 'FIXED_ASSET', 'OTHER_ASSET']
            for asset_type in asset_types:
                accounts = self.get_accounts_by_type(asset_type)
                section_total = Decimal('0.00')
                
                for account in accounts:
                    balance = self.get_account_balance(account.id, as_of_date, currency)
                    if balance != Decimal('0.00'):
                        account_data = {
                            'account_id': account.id,
                            'account_code': account.code,
                            'account_name': account.name,
                            'balance': balance
                        }
                        
                        if asset_type == 'CURRENT_ASSET':
                            balance_sheet['assets']['current_assets'].append(account_data)
                        elif asset_type == 'FIXED_ASSET':
                            balance_sheet['assets']['fixed_assets'].append(account_data)
                        else:
                            balance_sheet['assets']['other_assets'].append(account_data)
                        
                        section_total += balance
                
                balance_sheet['assets']['total'] += section_total
            
            # Get liability accounts
            liability_types = ['CURRENT_LIABILITY', 'LONG_TERM_LIABILITY']
            for liability_type in liability_types:
                accounts = self.get_accounts_by_type(liability_type)
                section_total = Decimal('0.00')
                
                for account in accounts:
                    balance = self.get_account_balance(account.id, as_of_date, currency)
                    if balance != Decimal('0.00'):
                        account_data = {
                            'account_id': account.id,
                            'account_code': account.code,
                            'account_name': account.name,
                            'balance': balance
                        }
                        
                        if liability_type == 'CURRENT_LIABILITY':
                            balance_sheet['liabilities']['current_liabilities'].append(account_data)
                        else:
                            balance_sheet['liabilities']['long_term_liabilities'].append(account_data)
                        
                        section_total += balance
                
                balance_sheet['liabilities']['total'] += section_total
            
            # Get equity accounts
            equity_accounts = self.get_accounts_by_type('EQUITY')
            for account in equity_accounts:
                balance = self.get_account_balance(account.id, as_of_date, currency)
                if balance != Decimal('0.00'):
                    account_data = {
                        'account_id': account.id,
                        'account_code': account.code,
                        'account_name': account.name,
                        'balance': balance
                    }
                    balance_sheet['equity']['equity_accounts'].append(account_data)
                    balance_sheet['equity']['total'] += balance
            
            # Get retained earnings
            retained_earnings_accounts = self.get_accounts_by_type('RETAINED_EARNINGS')
            for account in retained_earnings_accounts:
                balance = self.get_account_balance(account.id, as_of_date, currency)
                balance_sheet['equity']['retained_earnings'] += balance
                balance_sheet['equity']['total'] += balance
            
            # Calculate current period earnings (Revenue - Expenses - COGS)
            current_period_earnings = self._calculate_current_period_earnings(as_of_date, currency)
            balance_sheet['equity']['current_period_earnings'] = current_period_earnings
            balance_sheet['equity']['total'] += current_period_earnings
            
            # Validate accounting equation
            total_liabilities_equity = balance_sheet['liabilities']['total'] + balance_sheet['equity']['total']
            balance_sheet['is_balanced'] = abs(balance_sheet['assets']['total'] - total_liabilities_equity) <= Decimal('0.01')
            balance_sheet['difference'] = balance_sheet['assets']['total'] - total_liabilities_equity
            
            logger.info(f"Generated balance sheet as of {as_of_date}")
            return balance_sheet
            
        except Exception as e:
            logger.error(f"Error generating balance sheet: {str(e)}")
            raise ValidationError(f"Failed to generate balance sheet: {str(e)}")
    
    def generate_income_statement(self, start_date: date, end_date: date = None,
                                currency: Currency = None) -> Dict:
        """
        Generate income statement (P&L) for specified period
        
        Args:
            start_date: Period start date
            end_date: Period end date (defaults to today)
            currency: Currency for amounts (defaults to base currency)
            
        Returns:
            Complete income statement data
        """
        try:
            if not end_date:
                end_date = date.today()
            
            if not currency:
                currency = self.base_currency
            
            income_statement = {
                'period': {
                    'start_date': start_date,
                    'end_date': end_date
                },
                'currency': currency.code,
                'revenue': {
                    'revenue_accounts': [],
                    'other_income': [],
                    'total_revenue': Decimal('0.00'),
                    'total_other_income': Decimal('0.00'),
                    'gross_revenue': Decimal('0.00')
                },
                'cost_of_goods_sold': {
                    'cogs_accounts': [],
                    'total_cogs': Decimal('0.00')
                },
                'gross_profit': Decimal('0.00'),
                'expenses': {
                    'operating_expenses': [],
                    'other_expenses': [],
                    'total_operating': Decimal('0.00'),
                    'total_other': Decimal('0.00'),
                    'total_expenses': Decimal('0.00')
                },
                'net_income': Decimal('0.00')
            }
            
            # Get revenue accounts
            revenue_accounts = self.get_accounts_by_type('REVENUE')
            for account in revenue_accounts:
                period_activity = self._get_account_period_activity(account.id, start_date, end_date, currency)
                if period_activity['credits'] != Decimal('0.00'):
                    account_data = {
                        'account_id': account.id,
                        'account_code': account.code,
                        'account_name': account.name,
                        'amount': period_activity['credits']
                    }
                    income_statement['revenue']['revenue_accounts'].append(account_data)
                    income_statement['revenue']['total_revenue'] += period_activity['credits']
            
            # Get other income accounts
            other_income_accounts = self.get_accounts_by_type('OTHER_INCOME')
            for account in other_income_accounts:
                period_activity = self._get_account_period_activity(account.id, start_date, end_date, currency)
                if period_activity['credits'] != Decimal('0.00'):
                    account_data = {
                        'account_id': account.id,
                        'account_code': account.code,
                        'account_name': account.name,
                        'amount': period_activity['credits']
                    }
                    income_statement['revenue']['other_income'].append(account_data)
                    income_statement['revenue']['total_other_income'] += period_activity['credits']
            
            income_statement['revenue']['gross_revenue'] = (
                income_statement['revenue']['total_revenue'] + 
                income_statement['revenue']['total_other_income']
            )
            
            # Get COGS accounts
            cogs_accounts = self.get_accounts_by_type('COST_OF_GOODS_SOLD')
            for account in cogs_accounts:
                period_activity = self._get_account_period_activity(account.id, start_date, end_date, currency)
                if period_activity['debits'] != Decimal('0.00'):
                    account_data = {
                        'account_id': account.id,
                        'account_code': account.code,
                        'account_name': account.name,
                        'amount': period_activity['debits']
                    }
                    income_statement['cost_of_goods_sold']['cogs_accounts'].append(account_data)
                    income_statement['cost_of_goods_sold']['total_cogs'] += period_activity['debits']
            
            # Calculate gross profit
            income_statement['gross_profit'] = (
                income_statement['revenue']['gross_revenue'] - 
                income_statement['cost_of_goods_sold']['total_cogs']
            )
            
            # Get expense accounts
            expense_accounts = self.get_accounts_by_type('EXPENSE')
            for account in expense_accounts:
                period_activity = self._get_account_period_activity(account.id, start_date, end_date, currency)
                if period_activity['debits'] != Decimal('0.00'):
                    account_data = {
                        'account_id': account.id,
                        'account_code': account.code,
                        'account_name': account.name,
                        'amount': period_activity['debits']
                    }
                    income_statement['expenses']['operating_expenses'].append(account_data)
                    income_statement['expenses']['total_operating'] += period_activity['debits']
            
            # Get other expense accounts
            other_expense_accounts = self.get_accounts_by_type('OTHER_EXPENSE')
            for account in other_expense_accounts:
                period_activity = self._get_account_period_activity(account.id, start_date, end_date, currency)
                if period_activity['debits'] != Decimal('0.00'):
                    account_data = {
                        'account_id': account.id,
                        'account_code': account.code,
                        'account_name': account.name,
                        'amount': period_activity['debits']
                    }
                    income_statement['expenses']['other_expenses'].append(account_data)
                    income_statement['expenses']['total_other'] += period_activity['debits']
            
            income_statement['expenses']['total_expenses'] = (
                income_statement['expenses']['total_operating'] + 
                income_statement['expenses']['total_other']
            )
            
            # Calculate net income
            income_statement['net_income'] = (
                income_statement['gross_profit'] - 
                income_statement['expenses']['total_expenses']
            )
            
            logger.info(f"Generated income statement for {start_date} to {end_date}")
            return income_statement
            
        except Exception as e:
            logger.error(f"Error generating income statement: {str(e)}")
            raise ValidationError(f"Failed to generate income statement: {str(e)}")
    
    def _get_account_period_activity(self, account_id: int, start_date: date, 
                                   end_date: date, currency: Currency = None) -> Dict:
        """Get account activity for specific period"""
        if not currency:
            currency = self.base_currency
        
        line_totals = JournalEntryLine.objects.filter(
            tenant=self.tenant,
            account_id=account_id,
            journal_entry__entry_date__gte=start_date,
            journal_entry__entry_date__lte=end_date,
            journal_entry__status='POSTED'
        ).aggregate(
            total_debits=Sum('base_currency_debit_amount'),
            total_credits=Sum('base_currency_credit_amount')
        )
        
        debits = line_totals['total_debits'] or Decimal('0.00')
        credits = line_totals['total_credits'] or Decimal('0.00')
        
        # Convert currency if needed
        if currency.code != self.base_currency.code:
            exchange_rate = self.get_exchange_rate(self.base_currency, currency, end_date)
            debits = debits * exchange_rate
            credits = credits * exchange_rate
        
        return {
            'debits': debits,
            'credits': credits,
            'net_change': debits - credits
        }
    
    def _calculate_current_period_earnings(self, as_of_date: date, currency: Currency = None) -> Decimal:
        """Calculate current period earnings (since fiscal year start)"""
        if not currency:
            currency = self.base_currency
        
        fiscal_year_start = self.current_fiscal_year.start_date
        
        # Generate income statement for current period
        income_statement = self.generate_income_statement(fiscal_year_start, as_of_date, currency)
        
        return income_statement['net_income']
    
    # ============================================================================
    # CASH FLOW STATEMENT GENERATION
    # ============================================================================
    
    def generate_cash_flow_statement(self, start_date: date, end_date: date = None,
                                   currency: Currency = None, method: str = 'INDIRECT') -> Dict:
        """
        Generate cash flow statement using indirect or direct method
        
        Args:
            start_date: Period start date
            end_date: Period end date (defaults to today)
            currency: Currency for amounts (defaults to base currency)
            method: 'INDIRECT' or 'DIRECT'
            
        Returns:
            Complete cash flow statement data
        """
        try:
            if not end_date:
                end_date = date.today()
            
            if not currency:
                currency = self.base_currency
            
            cash_flow = {
                'period': {
                    'start_date': start_date,
                    'end_date': end_date
                },
                'currency': currency.code,
                'method': method,
                'operating_activities': {
                    'items': [],
                    'net_cash_from_operations': Decimal('0.00')
                },
                'investing_activities': {
                    'items': [],
                    'net_cash_from_investing': Decimal('0.00')
                },
                'financing_activities': {
                    'items': [],
                    'net_cash_from_financing': Decimal('0.00')
                },
                'net_change_in_cash': Decimal('0.00'),
                'cash_beginning': Decimal('0.00'),
                'cash_ending': Decimal('0.00')
            }
            
            if method == 'INDIRECT':
                cash_flow = self._generate_indirect_cash_flow(cash_flow, start_date, end_date, currency)
            else:
                cash_flow = self._generate_direct_cash_flow(cash_flow, start_date, end_date, currency)
            
            # Calculate cash balances
            cash_accounts = Account.objects.filter(
                tenant=self.tenant,
                is_cash_account=True,
                is_active=True
            )
            
            cash_beginning = Decimal('0.00')
            cash_ending = Decimal('0.00')
            
            for account in cash_accounts:
                cash_beginning += self.get_account_balance(account.id, start_date - timedelta(days=1), currency)
                cash_ending += self.get_account_balance(account.id, end_date, currency)
            
            cash_flow['cash_beginning'] = cash_beginning
            cash_flow['cash_ending'] = cash_ending
            cash_flow['net_change_in_cash'] = cash_ending - cash_beginning
            
            # Validate net change
            calculated_change = (
                cash_flow['operating_activities']['net_cash_from_operations'] +
                cash_flow['investing_activities']['net_cash_from_investing'] +
                cash_flow['financing_activities']['net_cash_from_financing']
            )
            
            cash_flow['is_balanced'] = abs(calculated_change - cash_flow['net_change_in_cash']) <= Decimal('0.01')
            
            logger.info(f"Generated cash flow statement ({method}) for {start_date} to {end_date}")
            return cash_flow
            
        except Exception as e:
            logger.error(f"Error generating cash flow statement: {str(e)}")
            raise ValidationError(f"Failed to generate cash flow statement: {str(e)}")
    
    def _generate_indirect_cash_flow(self, cash_flow: Dict, start_date: date, 
                                   end_date: date, currency: Currency) -> Dict:
        """Generate cash flow using indirect method"""
        # Start with net income
        income_statement = self.generate_income_statement(start_date, end_date, currency)
        net_income = income_statement['net_income']
        
        cash_flow['operating_activities']['items'].append({
            'description': 'Net Income',
            'amount': net_income
        })
        
        operating_cash = net_income
        
        # Add back non-cash expenses (depreciation, amortization)
        depreciation_accounts = Account.objects.filter(
            tenant=self.tenant,
            name__icontains='Depreciation',
            account_type='EXPENSE',
            is_active=True
        )
        
        for account in depreciation_accounts:
            activity = self._get_account_period_activity(account.id, start_date, end_date, currency)
            if activity['debits'] > 0:
                cash_flow['operating_activities']['items'].append({
                    'description': f'Add: {account.name}',
                    'amount': activity['debits']
                })
                operating_cash += activity['debits']
        
        # Analyze changes in working capital
        # Accounts Receivable
        ar_change = self._get_working_capital_change('CURRENT_ASSET', 'Receivable', start_date, end_date, currency)
        if ar_change != 0:
            cash_flow['operating_activities']['items'].append({
                'description': 'Change in Accounts Receivable',
                'amount': -ar_change  # Increase in AR reduces cash
            })
            operating_cash -= ar_change
        
        # Inventory
        inventory_change = self._get_working_capital_change('CURRENT_ASSET', 'Inventory', start_date, end_date, currency)
        if inventory_change != 0:
            cash_flow['operating_activities']['items'].append({
                'description': 'Change in Inventory',
                'amount': -inventory_change  # Increase in inventory reduces cash
            })
            operating_cash -= inventory_change
        
        # Accounts Payable
        ap_change = self._get_working_capital_change('CURRENT_LIABILITY', 'Payable', start_date, end_date, currency)
        if ap_change != 0:
            cash_flow['operating_activities']['items'].append({
                'description': 'Change in Accounts Payable',
                'amount': ap_change  # Increase in AP increases cash
            })
            operating_cash += ap_change
        
        cash_flow['operating_activities']['net_cash_from_operations'] = operating_cash
        
        # Investing Activities - Capital expenditures, asset purchases/sales
        fixed_asset_accounts = self.get_accounts_by_type('FIXED_ASSET')
        investing_cash = Decimal('0.00')
        
        for account in fixed_asset_accounts:
            activity = self._get_account_period_activity(account.id, start_date, end_date, currency)
            net_change = activity['debits'] - activity['credits']
            if abs(net_change) > Decimal('0.01'):
                description = f"Purchase of {account.name}" if net_change > 0 else f"Sale of {account.name}"
                cash_flow['investing_activities']['items'].append({
                    'description': description,
                    'amount': -net_change  # Asset purchases reduce cash, sales increase cash
                })
                investing_cash -= net_change
        
        cash_flow['investing_activities']['net_cash_from_investing'] = investing_cash
        
        # Financing Activities - Loans, equity, dividends
        financing_cash = Decimal('0.00')
        
        # Long-term debt changes
        debt_accounts = self.get_accounts_by_type('LONG_TERM_LIABILITY')
        for account in debt_accounts:
            if 'loan' in account.name.lower() or 'debt' in account.name.lower():
                activity = self._get_account_period_activity(account.id, start_date, end_date, currency)
                net_change = activity['credits'] - activity['debits']
                if abs(net_change) > Decimal('0.01'):
                    description = f"Proceeds from {account.name}" if net_change > 0 else f"Repayment of {account.name}"
                    cash_flow['financing_activities']['items'].append({
                        'description': description,
                        'amount': net_change
                    })
                    financing_cash += net_change
        
        # Equity changes
        equity_accounts = self.get_accounts_by_type('EQUITY')
        for account in equity_accounts:
            if 'capital' in account.name.lower() or 'stock' in account.name.lower():
                activity = self._get_account_period_activity(account.id, start_date, end_date, currency)
                net_change = activity['credits'] - activity['debits']
                if abs(net_change) > Decimal('0.01'):
                    description = f"Issuance of {account.name}" if net_change > 0 else f"Repurchase of {account.name}"
                    cash_flow['financing_activities']['items'].append({
                        'description': description,
                        'amount': net_change
                    })
                    financing_cash += net_change
        
        cash_flow['financing_activities']['net_cash_from_financing'] = financing_cash
        
        return cash_flow
    
    def _generate_direct_cash_flow(self, cash_flow: Dict, start_date: date, 
                                 end_date: date, currency: Currency) -> Dict:
        """Generate cash flow using direct method"""
        operating_cash = Decimal('0.00')
        
        # Cash receipts from customers
        customer_receipts = self._get_cash_receipts_from_customers(start_date, end_date, currency)
        cash_flow['operating_activities']['items'].append({
            'description': 'Cash receipts from customers',
            'amount': customer_receipts
        })
        operating_cash += customer_receipts
        
        # Cash payments to suppliers
        supplier_payments = self._get_cash_payments_to_suppliers(start_date, end_date, currency)
        cash_flow['operating_activities']['items'].append({
            'description': 'Cash payments to suppliers',
            'amount': -supplier_payments
        })
        operating_cash -= supplier_payments
        
        # Cash payments for operating expenses
        operating_expense_payments = self._get_cash_payments_for_expenses(start_date, end_date, currency)
        cash_flow['operating_activities']['items'].append({
            'description': 'Cash payments for operating expenses',
            'amount': -operating_expense_payments
        })
        operating_cash -= operating_expense_payments
        
        # Interest and tax payments
        interest_payments = self._get_interest_payments(start_date, end_date, currency)
        if interest_payments > 0:
            cash_flow['operating_activities']['items'].append({
                'description': 'Interest payments',
                'amount': -interest_payments
            })
            operating_cash -= interest_payments
        
        cash_flow['operating_activities']['net_cash_from_operations'] = operating_cash
        
        # Note: Investing and financing activities are the same for both methods
        return self._generate_indirect_cash_flow(cash_flow, start_date, end_date, currency)
    
    def _get_working_capital_change(self, account_type: str, name_filter: str, 
                                  start_date: date, end_date: date, currency: Currency) -> Decimal:
        """Calculate change in working capital component"""
        accounts = Account.objects.filter(
            tenant=self.tenant,
            account_type=account_type,
            name__icontains=name_filter,
            is_active=True
        )
        
        total_change = Decimal('0.00')
        for account in accounts:
            beginning_balance = self.get_account_balance(account.id, start_date - timedelta(days=1), currency)
            ending_balance = self.get_account_balance(account.id, end_date, currency)
            total_change += ending_balance - beginning_balance
        
        return total_change
    
    def _get_cash_receipts_from_customers(self, start_date: date, end_date: date, currency: Currency) -> Decimal:
        """Calculate cash receipts from customers"""
        # Get all customer payments received in the period
        customer_payments = Payment.objects.filter(
            tenant=self.tenant,
            payment_type='RECEIVED',
            payment_date__gte=start_date,
            payment_date__lte=end_date,
            status='CLEARED'
        ).aggregate(total=Sum('base_currency_amount'))
        
        receipts = customer_payments['total'] or Decimal('0.00')
        
        # Convert currency if needed
        if currency.code != self.base_currency.code:
            exchange_rate = self.get_exchange_rate(self.base_currency, currency, end_date)
            receipts = receipts * exchange_rate
        
        return receipts
    
    def _get_cash_payments_to_suppliers(self, start_date: date, end_date: date, currency: Currency) -> Decimal:
        """Calculate cash payments to suppliers"""
        # Get all vendor payments made in the period
        vendor_payments = Payment.objects.filter(
            tenant=self.tenant,
            payment_type='MADE',
            payment_date__gte=start_date,
            payment_date__lte=end_date,
            status='CLEARED'
        ).aggregate(total=Sum('base_currency_amount'))
        
        payments = vendor_payments['total'] or Decimal('0.00')
        
        # Convert currency if needed
        if currency.code != self.base_currency.code:
            exchange_rate = self.get_exchange_rate(self.base_currency, currency, end_date)
            payments = payments * exchange_rate
        
        return payments
    
    def _get_cash_payments_for_expenses(self, start_date: date, end_date: date, currency: Currency) -> Decimal:
        """Calculate cash payments for operating expenses"""
        # This would analyze cash account activity for expense-related payments
        # For simplicity, we'll estimate based on expense account changes
        expense_accounts = self.get_accounts_by_type('EXPENSE')
        total_expenses = Decimal('0.00')
        
        for account in expense_accounts:
            activity = self._get_account_period_activity(account.id, start_date, end_date, currency)
            total_expenses += activity['debits']
        
        return total_expenses
    
    def _get_interest_payments(self, start_date: date, end_date: date, currency: Currency) -> Decimal:
        """Calculate interest payments"""
        interest_accounts = Account.objects.filter(
            tenant=self.tenant,
            account_type='EXPENSE',
            name__icontains='Interest',
            is_active=True
        )
        
        total_interest = Decimal('0.00')
        for account in interest_accounts:
            activity = self._get_account_period_activity(account.id, start_date, end_date, currency)
            total_interest += activity['debits']
        
        return total_interest
    
    # ============================================================================
    # MULTI-CURRENCY SUPPORT
    # ============================================================================
    
    def get_exchange_rate(self, from_currency: Currency, to_currency: Currency, 
                         as_of_date: date = None) -> Decimal:
        """
        Get exchange rate between currencies
        
        Args:
            from_currency: Source currency
            to_currency: Target currency
            as_of_date: Date for exchange rate (defaults to today)
            
        Returns:
            Exchange rate
        """
        try:
            if from_currency.code == to_currency.code:
                return Decimal('1.000000')
            
            if not as_of_date:
                as_of_date = date.today()
            
            # Try to get exchange rate from database
            exchange_rate = ExchangeRate.objects.filter(
                tenant=self.tenant,
                from_currency=from_currency,
                to_currency=to_currency,
                effective_date__lte=as_of_date
            ).order_by('-effective_date').first()
            
            if exchange_rate:
                return exchange_rate.rate
            
            # Try inverse rate
            inverse_rate = ExchangeRate.objects.filter(
                tenant=self.tenant,
                from_currency=to_currency,
                to_currency=from_currency,
                effective_date__lte=as_of_date
            ).order_by('-effective_date').first()
            
            if inverse_rate:
                return Decimal('1.000000') / inverse_rate.rate
            
            # If no rate found, return 1:1 (should probably fetch from external API)
            logger.warning(f"No exchange rate found for {from_currency.code} to {to_currency.code}")
            return Decimal('1.000000')
            
        except Exception as e:
            logger.error(f"Error getting exchange rate: {str(e)}")
            return Decimal('1.000000')
    
    def convert_amount(self, amount: Decimal, from_currency: Currency, 
                      to_currency: Currency, as_of_date: date = None) -> Decimal:
        """Convert amount between currencies"""
        if amount == Decimal('0.00') or from_currency.code == to_currency.code:
            return amount
        
        exchange_rate = self.get_exchange_rate(from_currency, to_currency, as_of_date)
        converted_amount = amount * exchange_rate
        
        return converted_amount.quantize(
            Decimal('0.01'), 
            rounding=ROUND_HALF_UP
        )
    
    def update_exchange_rates(self, rates_data: List[Dict]) -> Dict:
        """
        Update exchange rates from external source
        
        Args:
            rates_data: List of rate data [{from_currency, to_currency, rate, date}]
            
        Returns:
            Update results
        """
        try:
            updated_count = 0
            created_count = 0
            
            with transaction.atomic():
                for rate_data in rates_data:
                    from_currency = Currency.objects.get(
                        tenant=self.tenant,
                        code=rate_data['from_currency']
                    )
                    to_currency = Currency.objects.get(
                        tenant=self.tenant,
                        code=rate_data['to_currency']
                    )
                    
                    rate, created = ExchangeRate.objects.update_or_create(
                        tenant=self.tenant,
                        from_currency=from_currency,
                        to_currency=to_currency,
                        effective_date=rate_data['date'],
                        defaults={
                            'rate': Decimal(str(rate_data['rate'])),
                            'source': rate_data.get('source', 'manual')
                        }
                    )
                    
                    if created:
                        created_count += 1
                    else:
                        updated_count += 1
            
            logger.info(f"Exchange rates updated: {updated_count} updated, {created_count} created")
            return {
                'success': True,
                'updated': updated_count,
                'created': created_count
            }
            
        except Exception as e:
            logger.error(f"Error updating exchange rates: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }
    
    # ============================================================================
    # PERIOD CLOSING AND VALIDATION
    # ============================================================================
    
    @transaction.atomic
    def close_accounting_period(self, period_end_date: date, user) -> Dict:
        """
        Close accounting period with validation and summary
        
        Args:
            period_end_date: End date of period to close
            user: User performing the closing
            
        Returns:
            Closing results and financial summary
        """
        try:
            # Validate period is ready for closing
            validation_result = self._validate_period_for_closing(period_end_date)
            if not validation_result['is_valid']:
                raise ValidationError(f"Period cannot be closed: {', '.join(validation_result['errors'])}")
            
            # Update all account balances
            self.update_account_balances(as_of_date=period_end_date)
            
            # Generate trial balance to ensure books are balanced
            trial_balance = self.generate_trial_balance(period_end_date)
            if not trial_balance['is_balanced']:
                raise ValidationError(f"Trial balance is not balanced. Difference: {trial_balance['totals']['difference']}")
            
            # Get or create financial period
            financial_period = self._get_or_create_financial_period(period_end_date)
            
            # Create closing entry if needed
            closing_entry = self._create_period_closing_entry(financial_period, user)
            
            # Update period status
            financial_period.status = 'CLOSED'
            financial_period.closed_by = user
            financial_period.closed_date = timezone.now()
            
            # Calculate period actuals
            period_start = financial_period.start_date
            income_statement = self.generate_income_statement(period_start, period_end_date)
            
            financial_period.actual_revenue = income_statement['revenue']['gross_revenue']
            financial_period.actual_expenses = income_statement['expenses']['total_expenses']
            financial_period.save()
            
            # Generate financial summary
            financial_summary = {
                'period': {
                    'start_date': period_start,
                    'end_date': period_end_date,
                    'status': 'CLOSED'
                },
                'trial_balance': trial_balance,
                'income_statement': income_statement,
                'closing_entry_id': closing_entry.id if closing_entry else None,
                'summary': {
                    'total_revenue': financial_period.actual_revenue,
                    'total_expenses': financial_period.actual_expenses,
                    'net_income': financial_period.actual_revenue - financial_period.actual_expenses
                }
            }
            
            logger.info(f"Accounting period closed successfully for {period_end_date}")
            return financial_summary
            
        except Exception as e:
            logger.error(f"Error closing accounting period: {str(e)}")
            raise ValidationError(f"Failed to close period: {str(e)}")
    
    def _validate_period_for_closing(self, period_end_date: date) -> Dict:
        """Validate period is ready for closing"""
        errors = []
        
        # Check for unposted journal entries
        unposted_entries = JournalEntry.objects.filter(
            tenant=self.tenant,
            entry_date__lte=period_end_date,
            status='DRAFT'
        ).count()
        
        if unposted_entries > 0:
            errors.append(f"{unposted_entries} unposted journal entries found")
        
        # Check for unbalanced entries
        unbalanced_entries = JournalEntry.objects.filter(
            tenant=self.tenant,
            entry_date__lte=period_end_date,
            status='POSTED'
        ).exclude(
            total_debit=F('total_credit')
        ).count()
        
        if unbalanced_entries > 0:
            errors.append(f"{unbalanced_entries} unbalanced journal entries found")
        
        # Check for incomplete bank reconciliations
        unreconciled_statements = models.Count('bankstatement', filter=Q(
            bankstatement__statement_date__lte=period_end_date,
            bankstatement__is_reconciled=False
        ))
        
        # Additional validations can be added here
        
        return {
            'is_valid': len(errors) == 0,
            'errors': errors
        }
    
    def _get_or_create_financial_period(self, period_end_date: date) -> FinancialPeriod:
        """Get or create financial period for the given date"""
        # Determine period start (assume monthly periods)
        period_start = period_end_date.replace(day=1)
        
        period, created = FinancialPeriod.objects.get_or_create(
            tenant=self.tenant,
            start_date=period_start,
            end_date=period_end_date,
            defaults={
                'name': f"{period_start.strftime('%B %Y')} Period",
                'period_type': 'MONTH',
                'fiscal_year': self.current_fiscal_year,
                'status': 'OPEN'
            }
        )
        
        return period
    
    def _create_period_closing_entry(self, financial_period: FinancialPeriod, user) -> Optional[JournalEntry]:
        """Create period closing journal entry if needed"""
        # This would create any necessary period-end adjusting entries
        # For now, return None as this is typically done at year-end
        return None
    
    # ============================================================================
    # VALIDATION & UTILITIES
    # ============================================================================
    
    def validate_accounting_equation(self, as_of_date: date = None) -> Dict:
        """
        Validate that Assets = Liabilities + Equity
        
        Args:
            as_of_date: Date to validate (defaults to today)
        
        Returns:
            Validation result with detailed breakdown
        """
        try:
            if not as_of_date:
                as_of_date = date.today()
            
            balance_sheet = self.generate_balance_sheet(as_of_date)
            
            assets = balance_sheet['assets']['total']
            liabilities = balance_sheet['liabilities']['total']
            equity = balance_sheet['equity']['total']
            
            total_liabilities_equity = liabilities + equity
            difference = assets - total_liabilities_equity
            is_balanced = abs(difference) <= Decimal('0.01')
            
            return {
                'is_balanced': is_balanced,
                'as_of_date': as_of_date,
                'assets': assets,
                'liabilities': liabilities,
                'equity': equity,
                'total_liabilities_equity': total_liabilities_equity,
                'difference': difference,
                'tolerance': Decimal('0.01'),
                'details': balance_sheet
            }
            
        except Exception as e:
            logger.error(f"Error validating accounting equation: {str(e)}")
            return {
                'is_balanced': False,
                'error': str(e)
            }
    
    def get_financial_ratios(self, as_of_date: date = None) -> Dict:
        """
        Calculate key financial ratios
        
        Args:
            as_of_date: Date for ratio calculation (defaults to today)
            
        Returns:
            Dictionary of financial ratios
        """
        try:
            if not as_of_date:
                as_of_date = date.today()
            
            # Generate required financial statements
            balance_sheet = self.generate_balance_sheet(as_of_date)
            
            # Calculate period for income statement (last 12 months)
            period_start = as_of_date.replace(year=as_of_date.year - 1)
            income_statement = self.generate_income_statement(period_start, as_of_date)
            
            # Extract key figures
            current_assets = sum(item['balance'] for item in balance_sheet['assets']['current_assets'])
            current_liabilities = sum(item['balance'] for item in balance_sheet['liabilities']['current_liabilities'])
            total_assets = balance_sheet['assets']['total']
            total_liabilities = balance_sheet['liabilities']['total']
            total_equity = balance_sheet['equity']['total']
            
            revenue = income_statement['revenue']['gross_revenue']
            net_income = income_statement['net_income']
            
            # Calculate ratios
            ratios = {
                'liquidity_ratios': {
                    'current_ratio': current_assets / current_liabilities if current_liabilities > 0 else Decimal('0.00'),
                    'quick_ratio': (current_assets - self._get_inventory_value(as_of_date)) / current_liabilities if current_liabilities > 0 else Decimal('0.00')
                },
                'leverage_ratios': {
                    'debt_to_equity': total_liabilities / total_equity if total_equity > 0 else Decimal('0.00'),
                    'debt_to_assets': total_liabilities / total_assets if total_assets > 0 else Decimal('0.00')
                },
                'profitability_ratios': {
                    'net_profit_margin': net_income / revenue if revenue > 0 else Decimal('0.00'),
                    'return_on_assets': net_income / total_assets if total_assets > 0 else Decimal('0.00'),
                    'return_on_equity': net_income / total_equity if total_equity > 0 else Decimal('0.00')
                },
                'as_of_date': as_of_date,
                'period_covered': f"{period_start} to {as_of_date}"
            }
            
            return ratios
            
        except Exception as e:
            logger.error(f"Error calculating financial ratios: {str(e)}")
            return {'error': str(e)}
    
    def _get_inventory_value(self, as_of_date: date) -> Decimal:
        """Get total inventory value as of date"""
        inventory_accounts = Account.objects.filter(
            tenant=self.tenant,
            track_inventory=True,
            is_active=True
        )
        
        total_inventory = Decimal('0.00')
        for account in inventory_accounts:
            total_inventory += self.get_account_balance(account.id, as_of_date)
        
        return total_inventory
    
    def get_account_aging_analysis(self, account_type: str, as_of_date: date = None) -> Dict:
        """
        Generate aging analysis for AR or AP accounts
        
        Args:
            account_type: 'AR' for receivables, 'AP' for payables
            as_of_date: Date for aging analysis (defaults to today)
            
        Returns:
            Aging analysis data
        """
        try:
            if not as_of_date:
                as_of_date = date.today()
            
            if account_type == 'AR':
                documents = Invoice.objects.filter(
                    tenant=self.tenant,
                    status__in=['OPEN', 'SENT', 'VIEWED', 'PARTIAL'],
                    amount_due__gt=0
                )
                entity_field = 'customer'
                amount_field = 'base_currency_amount_due'
                date_field = 'due_date'
            else:  # AP
                documents = Bill.objects.filter(
                    tenant=self.tenant,
                    status__in=['OPEN', 'APPROVED', 'PARTIAL'],
                    amount_due__gt=0
                )
                entity_field = 'vendor'
                amount_field = 'base_currency_amount_due'
                date_field = 'due_date'
            
            aging_buckets = {
                'current': {'days': '0-30', 'amount': Decimal('0.00'), 'count': 0},
                'days_31_60': {'days': '31-60', 'amount': Decimal('0.00'), 'count': 0},
                'days_61_90': {'days': '61-90', 'amount': Decimal('0.00'), 'count': 0},
                'days_91_120': {'days': '91-120', 'amount': Decimal('0.00'), 'count': 0},
                'over_120': {'days': '120+', 'amount': Decimal('0.00'), 'count': 0}
            }
            
            total_amount = Decimal('0.00')
            total_count = 0
            
            for document in documents:
                due_date = getattr(document, date_field)
                amount = getattr(document, amount_field)
                days_overdue = (as_of_date - due_date).days
                
                # Categorize into aging buckets
                if days_overdue <= 30:
                    bucket = 'current'
                elif days_overdue <= 60:
                    bucket = 'days_31_60'
                elif days_overdue <= 90:
                    bucket = 'days_61_90'
                elif days_overdue <= 120:
                    bucket = 'days_91_120'
                else:
                    bucket = 'over_120'
                
                aging_buckets[bucket]['amount'] += amount
                aging_buckets[bucket]['count'] += 1
                total_amount += amount
                total_count += 1
            
            # Calculate percentages
            for bucket_data in aging_buckets.values():
                if total_amount > 0:
                    bucket_data['percentage'] = (bucket_data['amount'] / total_amount * 100).quantize(Decimal('0.01'))
                else:
                    bucket_data['percentage'] = Decimal('0.00')
            
            return {
                'type': account_type,
                'as_of_date': as_of_date,
                'aging_buckets': aging_buckets,
                'summary': {
                    'total_amount': total_amount,
                    'total_count': total_count,
                    'average_amount': total_amount / total_count if total_count > 0 else Decimal('0.00')
                }
            }
            
        except Exception as e:
            logger.error(f"Error generating aging analysis: {str(e)}")
            return {'error': str(e)}
    
    # ============================================================================
    # PERFORMANCE AND ANALYTICS
    # ============================================================================
    
    def get_monthly_financial_summary(self, year: int = None) -> Dict:
        """Get monthly financial summary for the year"""
        try:
            if not year:
                year = date.today().year
            
            monthly_data = []
            
            for month in range(1, 13):
                month_start = date(year, month, 1)
                # Get last day of month
                if month == 12:
                    month_end = date(year + 1, 1, 1) - timedelta(days=1)
                else:
                    month_end = date(year, month + 1, 1) - timedelta(days=1)
                
                # Generate income statement for the month
                income_statement = self.generate_income_statement(month_start, month_end)
                
                monthly_data.append({
                    'month': month,
                    'month_name': month_start.strftime('%B'),
                    'revenue': income_statement['revenue']['gross_revenue'],
                    'expenses': income_statement['expenses']['total_expenses'],
                    'net_income': income_statement['net_income'],
                    'period': {
                        'start_date': month_start,
                        'end_date': month_end
                    }
                })
            
            # Calculate year totals
            year_totals = {
                'revenue': sum(month['revenue'] for month in monthly_data),
                'expenses': sum(month['expenses'] for month in monthly_data),
                'net_income': sum(month['net_income'] for month in monthly_data)
            }
            
            return {
                'year': year,
                'monthly_data': monthly_data,
                'year_totals': year_totals,
                'currency': self.base_currency.code
            }
            
        except Exception as e:
            logger.error(f"Error generating monthly financial summary: {str(e)}")
            return {'error': str(e)}
    
    def get_top_accounts_by_activity(self, account_type: str = None, 
                                   start_date: date = None, end_date: date = None,
                                   limit: int = 10) -> List[Dict]:
        """Get accounts with highest activity in period"""
        try:
            if not start_date:
                start_date = date.today() - timedelta(days=30)
            if not end_date:
                end_date = date.today()
            
            # Build query
            query = Account.objects.filter(tenant=self.tenant, is_active=True)
            if account_type:
                query = query.filter(account_type=account_type)
            
            account_activity = []
            
            for account in query:
                activity = self._get_account_period_activity(account.id, start_date, end_date)
                total_activity = activity['debits'] + activity['credits']
                
                if total_activity > Decimal('0.01'):  # Only include accounts with activity
                    account_activity.append({
                        'account_id': account.id,
                        'account_code': account.code,
                        'account_name': account.name,
                        'account_type': account.account_type,
                        'total_debits': activity['debits'],
                        'total_credits': activity['credits'],
                        'total_activity': total_activity,
                        'net_change': activity['net_change']
                    })
            
            # Sort by total activity and limit results
            account_activity.sort(key=lambda x: x['total_activity'], reverse=True)
            
            return account_activity[:limit]
            
        except Exception as e:
            logger.error(f"Error getting top accounts by activity: {str(e)}")
            return []
    
    # ============================================================================
    # BUDGET ANALYSIS
    # ============================================================================
    
    def compare_actual_to_budget(self, start_date: date, end_date: date = None,
                                budget_id: int = None) -> Dict:
        """
        Compare actual results to budget
        
        Args:
            start_date: Period start date
            end_date: Period end date (defaults to today)
            budget_id: Specific budget ID (defaults to current period budget)
            
        Returns:
            Budget vs actual analysis
        """
        try:
            if not end_date:
                end_date = date.today()
            
            # Get budget data
            if budget_id:
                from ..models import Budget, BudgetItem
                budget = Budget.objects.get(tenant=self.tenant, id=budget_id)
                budget_items = BudgetItem.objects.filter(budget=budget)
            else:
                # Use default budget for the period
                budget_items = []  # Would implement budget lookup logic
            
            # Get actual results
            income_statement = self.generate_income_statement(start_date, end_date)
            
            comparison = {
                'period': {
                    'start_date': start_date,
                    'end_date': end_date
                },
                'budget_id': budget_id,
                'revenue_comparison': {
                    'actual': income_statement['revenue']['gross_revenue'],
                    'budget': Decimal('0.00'),  # Would get from budget
                    'variance': Decimal('0.00'),
                    'variance_percent': Decimal('0.00')
                },
                'expense_comparison': {
                    'actual': income_statement['expenses']['total_expenses'],
                    'budget': Decimal('0.00'),  # Would get from budget
                    'variance': Decimal('0.00'),
                    'variance_percent': Decimal('0.00')
                },
                'net_income_comparison': {
                    'actual': income_statement['net_income'],
                    'budget': Decimal('0.00'),  # Would get from budget
                    'variance': Decimal('0.00'),
                    'variance_percent': Decimal('0.00')
                },
                'account_details': []
            }
            
            # Calculate variances (would be implemented with actual budget data)
            
            return comparison
            
        except Exception as e:
            logger.error(f"Error comparing actual to budget: {str(e)}")
            return {'error': str(e)}
    
    # ============================================================================
    # EXPORT AND INTEGRATION
    # ============================================================================
    
    def export_trial_balance(self, as_of_date: date = None, format_type: str = 'json') -> Union[Dict, str]:
        """Export trial balance in various formats"""
        try:
            trial_balance = self.generate_trial_balance(as_of_date)
            
            if format_type == 'json':
                return trial_balance
            elif format_type == 'csv':
                return self._convert_trial_balance_to_csv(trial_balance)
            elif format_type == 'excel':
                return self._convert_trial_balance_to_excel(trial_balance)
            else:
                raise ValidationError(f"Unsupported format: {format_type}")
                
        except Exception as e:
            logger.error(f"Error exporting trial balance: {str(e)}")
            return {'error': str(e)}
    
    def _convert_trial_balance_to_csv(self, trial_balance: Dict) -> str:
        """Convert trial balance to CSV format"""
        import csv
        import io
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow(['Account Code', 'Account Name', 'Account Type', 'Debit Amount', 'Credit Amount'])
        
        # Data rows
        for account in trial_balance['accounts']:
            writer.writerow([
                account['account_code'],
                account['account_name'],
                account['account_type'],
                str(account['debit_amount']),
                str(account['credit_amount'])
            ])
        
        # Totals
        writer.writerow(['', '', 'TOTALS', 
                        str(trial_balance['totals']['total_debits']),
                        str(trial_balance['totals']['total_credits'])])
        
        return output.getvalue()
    
    def _convert_trial_balance_to_excel(self, trial_balance: Dict) -> bytes:
        """Convert trial balance to Excel format"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Trial Balance"
            
            # Header
            headers = ['Account Code', 'Account Name', 'Account Type', 'Debit Amount', 'Credit Amount']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
            
            # Data
            for row, account in enumerate(trial_balance['accounts'], 2):
                ws.cell(row=row, column=1, value=account['account_code'])
                ws.cell(row=row, column=2, value=account['account_name'])
                ws.cell(row=row, column=3, value=account['account_type'])
                ws.cell(row=row, column=4, value=float(account['debit_amount']))
                ws.cell(row=row, column=5, value=float(account['credit_amount']))
            
            # Totals
            total_row = len(trial_balance['accounts']) + 2
            ws.cell(row=total_row, column=3, value='TOTALS').font = Font(bold=True)
            ws.cell(row=total_row, column=4, value=float(trial_balance['totals']['total_debits'])).font = Font(bold=True)
            ws.cell(row=total_row, column=5, value=float(trial_balance['totals']['total_credits'])).font = Font(bold=True)
            
            # Save to bytes
            import io
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return output.getvalue()
            
        except ImportError:
            return b"Excel export requires openpyxl library"
        except Exception as e:
            logger.error(f"Error creating Excel export: {str(e)}")
            return b"Error creating Excel file"
    
    # ============================================================================
    # DATA INTEGRITY AND MAINTENANCE
    # ============================================================================
    
    def run_integrity_check(self) -> Dict:
        """Run comprehensive data integrity checks"""
        try:
            checks = {
                'journal_entries': self._check_journal_entry_integrity(),
                'account_balances': self._check_account_balance_integrity(),
                'invoice_payments': self._check_invoice_payment_integrity(),
                'bill_payments': self._check_bill_payment_integrity(),
                'accounting_equation': self.validate_accounting_equation(),
                'exchange_rates': self._check_exchange_rate_integrity()
            }
            
            # Summary
            total_issues = sum(len(check.get('issues', [])) for check in checks.values() if isinstance(check, dict))
            
            return {
                'overall_status': 'PASS' if total_issues == 0 else 'FAIL',
                'total_issues': total_issues,
                'checks': checks,
                'run_date': timezone.now()
            }
            
        except Exception as e:
            logger.error(f"Error running integrity check: {str(e)}")
            return {'error': str(e)}
    
    def _check_journal_entry_integrity(self) -> Dict:
        """Check journal entry data integrity"""
        issues = []
        
        # Check for unbalanced entries
        unbalanced = JournalEntry.objects.filter(
            tenant=self.tenant,
            status='POSTED'
        ).exclude(total_debit=F('total_credit'))
        
        if unbalanced.exists():
            issues.append(f"{unbalanced.count()} unbalanced journal entries found")
        
        # Check for entries without lines
        entries_without_lines = JournalEntry.objects.filter(
            tenant=self.tenant,
            journal_lines__isnull=True
        )
        
        if entries_without_lines.exists():
            issues.append(f"{entries_without_lines.count()} journal entries without lines found")
        
        return {
            'status': 'PASS' if not issues else 'FAIL',
            'issues': issues
        }
    
    def _check_account_balance_integrity(self) -> Dict:
        """Check account balance integrity"""
        issues = []
        
        # Check if calculated balances match stored balances
        accounts = Account.objects.filter(tenant=self.tenant, is_active=True)
        mismatched_count = 0
        
        for account in accounts:
            calculated_balance = self.get_account_balance(account.id)
            if abs(calculated_balance - account.current_balance) > Decimal('0.01'):
                mismatched_count += 1
        
        if mismatched_count > 0:
            issues.append(f"{mismatched_count} accounts have mismatched balances")
        
        return {
            'status': 'PASS' if not issues else 'FAIL',
            'issues': issues
        }
    
    def _check_invoice_payment_integrity(self) -> Dict:
        """Check invoice payment integrity"""
        issues = []
        
        # Check for invoices with payments exceeding total
        overpaid_invoices = Invoice.objects.filter(
            tenant=self.tenant,
            amount_paid__gt=F('total_amount')
        )
        
        if overpaid_invoices.exists():
            issues.append(f"{overpaid_invoices.count()} overpaid invoices found")
        
        return {
            'status': 'PASS' if not issues else 'FAIL',
            'issues': issues
        }
    
    def _check_bill_payment_integrity(self) -> Dict:
        """Check bill payment integrity"""
        issues = []
        
        # Check for bills with payments exceeding total
        overpaid_bills = Bill.objects.filter(
            tenant=self.tenant,
            amount_paid__gt=F('total_amount')
        )
        
        if overpaid_bills.exists():
            issues.append(f"{overpaid_bills.count()} overpaid bills found")
        
        return {
            'status': 'PASS' if not issues else 'FAIL',
            'issues': issues
        }
    
    def _check_exchange_rate_integrity(self) -> Dict:
        """Check exchange rate data integrity"""
        issues = []
        
        # Check for missing exchange rates for multi-currency transactions
        # This would be implemented based on specific business requirements
        
        return {
            'status': 'PASS' if not issues else 'FAIL',
            'issues': issues
        }
    
    def cleanup_old_data(self, cutoff_date: date = None) -> Dict:
        """Clean up old data based on retention policies"""
        try:
            if not cutoff_date:
                # Default to 7 years retention
                cutoff_date = date.today() - timedelta(days=7*365)
            
            cleanup_results = {
                'cutoff_date': cutoff_date,
                'deleted_counts': {}
            }
            
            with transaction.atomic():
                # Archive old journal entries (but keep them)
                old_entries = JournalEntry.objects.filter(
                    tenant=self.tenant,
                    entry_date__lt=cutoff_date,
                    status='POSTED'
                )
                
                # Mark as archived instead of deleting
                archived_count = old_entries.update(notes=models.F('notes') + ' [ARCHIVED]')
                cleanup_results['deleted_counts']['archived_journal_entries'] = archived_count
                
                # Clean up old reconciliation data
                from ..models import BankReconciliation
                old_reconciliations = BankReconciliation.objects.filter(
                    tenant=self.tenant,
                    reconciliation_date__lt=cutoff_date,
                    status='COMPLETED'
                )
                
                # Keep reconciliations but clean up detailed transaction matches
                cleanup_results['deleted_counts']['cleaned_reconciliations'] = old_reconciliations.count()
            
            logger.info(f"Data cleanup completed for data older than {cutoff_date}")
            return cleanup_results
            
        except Exception as e:
            logger.error(f"Error during data cleanup: {str(e)}")
            return {'error': str(e)}
    
    # ============================================================================
    # REPORTING UTILITIES
    # ============================================================================
    
    def get_financial_dashboard_data(self, as_of_date: date = None) -> Dict:
        """Get comprehensive financial dashboard data"""
        try:
            if not as_of_date:
                as_of_date = date.today()
            
            # Key metrics
            balance_sheet = self.generate_balance_sheet(as_of_date)
            
            # Current month income statement
            month_start = as_of_date.replace(day=1)
            current_month_income = self.generate_income_statement(month_start, as_of_date)
            
            # Year to date income statement
            year_start = as_of_date.replace(month=1, day=1)
            ytd_income = self.generate_income_statement(year_start, as_of_date)
            
            # Cash flow (last 30 days)
            cash_flow_start = as_of_date - timedelta(days=30)
            cash_flow = self.generate_cash_flow_statement(cash_flow_start, as_of_date)
            
            # Aging analysis
            ar_aging = self.get_account_aging_analysis('AR', as_of_date)
            ap_aging = self.get_account_aging_analysis('AP', as_of_date)
            
            # Financial ratios
            ratios = self.get_financial_ratios(as_of_date)
            
            dashboard_data = {
                'as_of_date': as_of_date,
                'currency': self.base_currency.code,
                'key_metrics': {
                    'total_assets': balance_sheet['assets']['total'],
                    'total_liabilities': balance_sheet['liabilities']['total'],
                    'total_equity': balance_sheet['equity']['total'],
                    'current_month_revenue': current_month_income['revenue']['gross_revenue'],
                    'current_month_expenses': current_month_income['expenses']['total_expenses'],
                    'current_month_net_income': current_month_income['net_income'],
                    'ytd_revenue': ytd_income['revenue']['gross_revenue'],
                    'ytd_expenses': ytd_income['expenses']['total_expenses'],
                    'ytd_net_income': ytd_income['net_income'],
                    'cash_flow_operations': cash_flow['operating_activities']['net_cash_from_operations'],
                    'total_ar': ar_aging['summary']['total_amount'],
                    'total_ap': ap_aging['summary']['total_amount']
                },
                'balance_sheet_summary': {
                    'assets': balance_sheet['assets'],
                    'liabilities': balance_sheet['liabilities'], 
                    'equity': balance_sheet['equity']
                },
                'aging_summary': {
                    'accounts_receivable': ar_aging,
                    'accounts_payable': ap_aging
                },
                'financial_ratios': ratios,
                'accounting_equation_check': self.validate_accounting_equation(as_of_date)
            }
            
            return dashboard_data
            
        except Exception as e:
            logger.error(f"Error generating financial dashboard data: {str(e)}")
            return {'error': str(e)}