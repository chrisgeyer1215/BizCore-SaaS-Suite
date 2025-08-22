"""
Finance Exports Utilities
Data export utilities for financial data
"""

import csv
import json
from decimal import Decimal
from typing import List, Dict, Any
from django.http import HttpResponse
from django.db.models import QuerySet


class CSVExporter:
    """CSV export utilities"""
    
    @staticmethod
    def export_to_csv(data: List[Dict], filename: str, 
                     field_mappings: Dict[str, str] = None) -> HttpResponse:
        """Export data to CSV"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        
        if not data:
            return response
        
        # Get field names
        if field_mappings:
            fieldnames = list(field_mappings.keys())
        else:
            fieldnames = list(data[0].keys())
        
        writer = csv.DictWriter(response, fieldnames=fieldnames)
        writer.writeheader()
        
        # Write data rows
        for row in data:
            if field_mappings:
                # Map fields according to field_mappings
                mapped_row = {}
                for csv_field, data_field in field_mappings.items():
                    mapped_row[csv_field] = row.get(data_field, '')
                writer.writerow(mapped_row)
            else:
                writer.writerow(row)
        
        return response
    
    @staticmethod
    def export_accounts_to_csv(accounts: QuerySet, filename: str = 'chart_of_accounts') -> HttpResponse:
        """Export chart of accounts to CSV"""
        field_mappings = {
            'Account Code': 'code',
            'Account Name': 'name',
            'Account Type': 'account_type',
            'Category': 'category__name',
            'Normal Balance': 'normal_balance',
            'Current Balance': 'current_balance',
            'Currency': 'currency__code',
            'Status': 'is_active'
        }
        
        data = []
        for account in accounts:
            data.append({
                'code': account.code,
                'name': account.name,
                'account_type': account.get_account_type_display(),
                'category__name': account.category.name if account.category else '',
                'normal_balance': account.normal_balance,
                'current_balance': str(account.current_balance or 0),
                'currency__code': account.currency.code if account.currency else '',
                'is_active': 'Active' if account.is_active else 'Inactive'
            })
        
        return CSVExporter.export_to_csv(data, filename, field_mappings)


class JSONExporter:
    """JSON export utilities"""
    
    @staticmethod
    def export_to_json(data: Any, filename: str) -> HttpResponse:
        """Export data to JSON"""
        response = HttpResponse(
            json.dumps(data, indent=2, default=str),
            content_type='application/json'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.json"'
        return response


class ExcelExporter:
    """Excel export utilities"""
    
    @staticmethod
    def export_to_excel(data: List[Dict], filename: str, 
                       sheet_name: str = 'Data',
                       field_mappings: Dict[str, str] = None) -> HttpResponse:
        """Export data to Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            return HttpResponse("Excel export requires openpyxl library", status=400)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        if not data:
            return HttpResponse("No data to export", status=400)
        
        # Get field names
        if field_mappings:
            fieldnames = list(field_mappings.keys())
        else:
            fieldnames = list(data[0].keys())
        
        # Write headers
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for col, field in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col, value=field)
            cell.font = header_font
            cell.fill = header_fill
        
        # Write data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, field in enumerate(fieldnames, 1):
                if field_mappings:
                    # Map fields according to field_mappings
                    data_field = field_mappings[field]
                    value = row_data.get(data_field, '')
                else:
                    value = row_data.get(field, '')
                
                # Format value
                if isinstance(value, Decimal):
                    value = float(value)
                elif value is None:
                    value = ''
                
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        
        # Save to response
        wb.save(response)
        return response
