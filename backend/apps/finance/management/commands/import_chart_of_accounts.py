"""
Management command to import chart of accounts from CSV/Excel files
"""

import csv
import logging
from decimal import Decimal
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from apps.finance.models import Account, AccountCategory, AccountType
from apps.core.models import Tenant


logger = logging.getLogger(__name__)


class Command(BaseCommand):
    help = 'Import chart of accounts from CSV or Excel file'
    
    def add_arguments(self, parser):
        parser.add_argument(
            'tenant_schema',
            type=str,
            help='Tenant schema name to import accounts for'
        )
        parser.add_argument(
            'file_path',
            type=str,
            help='Path to CSV or Excel file'
        )
        parser.add_argument(
            '--file-format',
            choices=['csv', 'excel'],
            default='csv',
            help='File format (default: csv)'
        )
        parser.add_argument(
            '--update-existing',
            action='store_true',
            help='Update existing accounts if they exist'
        )
        parser.add_argument(
            '--create-categories',
            action='store_true',
            default=True,
            help='Automatically create missing categories'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Preview changes without saving'
        )
        parser.add_argument(
            '--delimiter',
            type=str,
            default=',',
            help='CSV delimiter (default: ,)'
        )
    
    def handle(self, *args, **options):
        tenant_schema = options['tenant_schema']
        file_path = options['file_path']
        file_format = options['file_format']
        update_existing = options['update_existing']
        create_categories = options['create_categories']
        dry_run = options['dry_run']
        delimiter = options['delimiter']
        
        # Validate tenant exists
        try:
            tenant = Tenant.objects.get(schema_name=tenant_schema)
        except Tenant.DoesNotExist:
            raise CommandError(f'Tenant with schema "{tenant_schema}" does not exist.')
        
        # Validate file exists
        try:
            with open(file_path, 'r') as f:
                pass
        except FileNotFoundError:
            raise CommandError(f'File "{file_path}" not found.')
        
        self.stdout.write(
            self.style.SUCCESS(f'Starting import for tenant: {tenant.name}')
        )
        
        if dry_run:
            self.stdout.write(
                self.style.WARNING('DRY RUN MODE - No changes will be saved')
            )
        
        # Parse file based on format
        if file_format == 'csv':
            accounts_data = self._parse_csv(file_path, delimiter)
        else:
            accounts_data = self._parse_excel(file_path)
        
        if not accounts_data:
            raise CommandError('No valid data found in file.')
        
        self.stdout.write(f'Found {len(accounts_data)} accounts to process')
        
        # Process accounts
        stats = self._process_accounts(
            tenant, accounts_data, update_existing, create_categories, dry_run
        )
        
        # Display results
        self._display_results(stats, dry_run)
    
    def _parse_csv(self, file_path, delimiter):
        """Parse CSV file and return list of account dictionaries"""
        accounts_data = []
        
        with open(file_path, 'r', encoding='utf-8') as file:
            reader = csv.DictReader(file, delimiter=delimiter)
            
            for row_num, row in enumerate(reader, start=2):
                try:
                    account_data = self._parse_csv_row(row, row_num)
                    if account_data:
                        accounts_data.append(account_data)
                except Exception as e:
                    self.stdout.write(
                        self.style.ERROR(f'Error parsing row {row_num}: {e}')
                    )
        
        return accounts_data
    
    def _parse_excel(self, file_path):
        """Parse Excel file and return list of account dictionaries"""
        accounts_data = []
        
        try:
            workbook = load_workbook(filename=file_path, read_only=True)
            worksheet = workbook.active
            
            # Get headers from first row
            headers = []
            for cell in worksheet[1]:
                headers.append(cell.value or f'Column_{len(headers)}')
            
            # Process data rows
            for row_num, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                try:
                    row_data = {}
                    for i, cell in enumerate(row):
                        if i < len(headers):
                            row_data[headers[i]] = cell.value
                    
                    account_data = self._parse_excel_row(row_data, row_num)
                    if account_data:
                        accounts_data.append(account_data)
                except Exception as e:
                    self.stdout.write(
                        self.style.ERROR(f'Error parsing row {row_num}: {e}')
                    )
            
            workbook.close()
            
        except Exception as e:
            raise CommandError(f'Error reading Excel file: {e}')
        
        return accounts_data
    
    def _parse_csv_row(self, row, row_num):
        """Parse a single CSV row into account data"""
        # Map common column names
        column_mapping = {
            'code': ['code', 'account_code', 'account_code', 'acct_code'],
            'name': ['name', 'account_name', 'description', 'acct_name'],
            'type': ['type', 'account_type', 'acct_type', 'category_type'],
            'category': ['category', 'account_category', 'acct_category', 'group'],
            'description': ['description', 'desc', 'notes', 'comment'],
            'normal_balance': ['normal_balance', 'balance_type', 'balance', 'debit_credit'],
            'opening_balance': ['opening_balance', 'opening_bal', 'start_balance', 'balance'],
            'currency': ['currency', 'curr', 'currency_code'],
            'is_active': ['is_active', 'active', 'status', 'enabled']
        }
        
        account_data = {}
        
        # Find and map columns
        for field, possible_names in column_mapping.items():
            for col_name in possible_names:
                if col_name in row and row[col_name]:
                    account_data[field] = str(row[col_name]).strip()
                    break
        
        # Validate required fields
        if not account_data.get('code'):
            self.stdout.write(
                self.style.WARNING(f'Row {row_num}: Missing account code, skipping')
            )
            return None
        
        if not account_data.get('name'):
            self.stdout.write(
                self.style.WARNING(f'Row {row_num}: Missing account name, skipping')
            )
            return None
        
        return account_data
    
    def _parse_excel_row(self, row_data, row_num):
        """Parse a single Excel row into account data"""
        # Convert Excel row to same format as CSV
        return self._parse_csv_row(row_data, row_num)
    
    def _process_accounts(self, tenant, accounts_data, update_existing, create_categories, dry_run):
        """Process accounts data and create/update accounts"""
        stats = {
            'created': 0,
            'updated': 0,
            'skipped': 0,
            'errors': 0,
            'categories_created': 0
        }
        
        with transaction.atomic():
            for account_data in accounts_data:
                try:
                    result = self._process_single_account(
                        tenant, account_data, update_existing, create_categories, dry_run
                    )
                    
                    if result == 'created':
                        stats['created'] += 1
                    elif result == 'updated':
                        stats['updated'] += 1
                    elif result == 'skipped':
                        stats['skipped'] += 1
                    else:
                        stats['errors'] += 1
                        
                except Exception as e:
                    stats['errors'] += 1
                    self.stdout.write(
                        self.style.ERROR(f'Error processing account {account_data.get("code", "Unknown")}: {e}')
                    )
        
        return stats
    
    def _process_single_account(self, tenant, account_data, update_existing, create_categories, dry_run):
        """Process a single account"""
        code = account_data['code']
        name = account_data['name']
        
        # Check if account already exists
        try:
            existing_account = Account.objects.get(tenant=tenant, code=code)
            if not update_existing:
                return 'skipped'
        except Account.DoesNotExist:
            existing_account = None
        
        # Process category
        category = None
        if account_data.get('category'):
            category = self._get_or_create_category(
                tenant, account_data['category'], account_data.get('type'), create_categories, dry_run
            )
        
        # Process account type
        account_type = self._parse_account_type(account_data.get('type'))
        
        # Process normal balance
        normal_balance = self._parse_normal_balance(account_data.get('normal_balance'))
        
        # Process opening balance
        opening_balance = self._parse_decimal(account_data.get('opening_balance'), '0.00')
        
        # Process currency
        currency = self._get_currency(tenant, account_data.get('currency'))
        
        # Process active status
        is_active = self._parse_boolean(account_data.get('is_active'), True)
        
        if dry_run:
            if existing_account:
                self.stdout.write(f'Would update: {code} - {name}')
            else:
                self.stdout.write(f'Would create: {code} - {name}')
            return 'created' if not existing_account else 'updated'
        
        # Create or update account
        if existing_account:
            existing_account.name = name
            existing_account.description = account_data.get('description', '')
            existing_account.category = category
            existing_account.account_type = account_type
            existing_account.normal_balance = normal_balance
            existing_account.opening_balance = opening_balance
            existing_account.opening_balance_date = timezone.now().date()
            existing_account.currency = currency
            existing_account.is_active = is_active
            existing_account.save()
            return 'updated'
        else:
            Account.objects.create(
                tenant=tenant,
                code=code,
                name=name,
                description=account_data.get('description', ''),
                category=category,
                account_type=account_type,
                normal_balance=normal_balance,
                opening_balance=opening_balance,
                opening_balance_date=timezone.now().date(),
                currency=currency,
                is_active=is_active
            )
            return 'created'
    
    def _get_or_create_category(self, tenant, category_name, category_type, create_categories, dry_run):
        """Get existing category or create new one"""
        try:
            return AccountCategory.objects.get(tenant=tenant, name=category_name)
        except AccountCategory.DoesNotExist:
            if not create_categories:
                return None
            
            # Parse account type for category
            account_type = self._parse_account_type(category_type)
            
            if dry_run:
                self.stdout.write(f'Would create category: {category_name}')
                return None
            
            category = AccountCategory.objects.create(
                tenant=tenant,
                name=category_name,
                code=category_name[:10].upper(),
                account_type=account_type,
                is_active=True
            )
            
            return category
    
    def _parse_account_type(self, type_str):
        """Parse account type string to AccountType choice"""
        if not type_str:
            return AccountType.ASSET
        
        type_str = str(type_str).upper().strip()
        
        # Map common variations
        type_mapping = {
            'ASSET': AccountType.ASSET,
            'ASSETS': AccountType.ASSET,
            'A': AccountType.ASSET,
            'LIABILITY': AccountType.LIABILITY,
            'LIABILITIES': AccountType.LIABILITY,
            'L': AccountType.LIABILITY,
            'EQUITY': AccountType.EQUITY,
            'E': AccountType.EQUITY,
            'REVENUE': AccountType.REVENUE,
            'REVENUES': AccountType.REVENUE,
            'R': AccountType.REVENUE,
            'EXPENSE': AccountType.EXPENSE,
            'EXPENSES': AccountType.EXPENSE,
            'X': AccountType.EXPENSE
        }
        
        return type_mapping.get(type_str, AccountType.ASSET)
    
    def _parse_normal_balance(self, balance_str):
        """Parse normal balance string"""
        if not balance_str:
            return 'DEBIT'
        
        balance_str = str(balance_str).upper().strip()
        
        if balance_str in ['CREDIT', 'CR', 'C']:
            return 'CREDIT'
        else:
            return 'DEBIT'
    
    def _parse_decimal(self, value, default='0.00'):
        """Parse decimal value"""
        if not value:
            return Decimal(default)
        
        try:
            return Decimal(str(value))
        except (ValueError, TypeError):
            return Decimal(default)
    
    def _get_currency(self, tenant, currency_code):
        """Get currency object"""
        if not currency_code:
            # Get default currency
            try:
                from apps.finance.models import Currency
                return Currency.objects.filter(tenant=tenant, is_default=True).first()
            except:
                return None
        
        try:
            from apps.finance.models import Currency
            return Currency.objects.get(tenant=tenant, code=currency_code.upper())
        except:
            return None
    
    def _parse_boolean(self, value, default=True):
        """Parse boolean value"""
        if value is None:
            return default
        
        if isinstance(value, bool):
            return value
        
        value_str = str(value).lower().strip()
        
        if value_str in ['true', '1', 'yes', 'y', 'active', 'enabled']:
            return True
        elif value_str in ['false', '0', 'no', 'n', 'inactive', 'disabled']:
            return False
        else:
            return default
    
    def _display_results(self, stats, dry_run):
        """Display import results"""
        self.stdout.write('\n' + '='*50)
        self.stdout.write('IMPORT RESULTS')
        self.stdout.write('='*50)
        
        if dry_run:
            self.stdout.write('DRY RUN MODE - No changes were saved')
            self.stdout.write('')
        
        self.stdout.write(f'Accounts created: {stats["created"]}')
        self.stdout.write(f'Accounts updated: {stats["updated"]}')
        self.stdout.write(f'Accounts skipped: {stats["skipped"]}')
        self.stdout.write(f'Categories created: {stats["categories_created"]}')
        self.stdout.write(f'Errors: {stats["errors"]}')
        
        total_processed = stats['created'] + stats['updated'] + stats['skipped']
        self.stdout.write(f'Total processed: {total_processed}')
        
        if stats['errors'] == 0:
            self.stdout.write(
                self.style.SUCCESS('Import completed successfully!')
            )
        else:
            self.stdout.write(
                self.style.WARNING(f'Import completed with {stats["errors"]} errors')
            )