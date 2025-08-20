import io
import csv
import json
import hashlib
from decimal import Decimal
from datetime import datetime, timedelta
from django.db import models
from django.db.models import Q, F, Sum, Count, Avg, Max, Min
from django.utils import timezone
from django.template.loader import render_to_string
from django.http import HttpResponse
from typing import List, Dict, Any, Optional, Union
from ..base import BaseService, ServiceResult
from ...models import (
    InventoryReport, ReportTemplate, ReportSchedule, ReportExecution,
    StockItem, Product, StockMovement, StockValuationLayer,
    PurchaseOrder, Supplier, Warehouse
)

try:
    import pandas as pd
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

class ReportService(BaseService):
    """
    Service for generating inventory reports
    """
    
    def generate_report(self, template_id: int, filters: Dict[str, Any] = None,
                       output_format: str = 'PDF') -> ServiceResult:
        """Generate a report based on template and filters"""
        try:
            template = ReportTemplate.objects.get(id=template_id, tenant=self.tenant)
            
            if not template.can_access(self.user):
                return ServiceResult.error("Access denied to this report template")
            
            # Create report record
            report = InventoryReport.objects.create(
                tenant=self.tenant,
                template=template,
                created_by=self.user,
                filters_applied=filters or {},
                output_format=output_format
            )
            
            try:
                report.start_generation()
                
                # Generate report data
                data_result = self._get_report_data(template, filters or {})
                if not data_result.is_success:
                    report.mark_failed(data_result.message)
                    return data_result
                
                # Generate file based on output format
                file_result = self._generate_report_file(
                    report, data_result.data, output_format
                )
                
                if file_result.is_success:
                    # Mark report as completed
                    file_info = file_result.data
                    report.mark_completed(
                        file_path=file_info['file_path'],
                        file_size=file_info['file_size'],
                        total_records=len(data_result.data.get('data', [])),
                        file_hash=file_info.get('file_hash')
                    )
                    
                    # Update template usage
                    template.mark_used()
                    
                    self.log_operation('generate_report', {
                        'report_id': report.id,
                        'template_name': template.name,
                        'output_format': output_format,
                        'total_records': report.total_records
                    })
                    
                    return ServiceResult.success(
                        data=report,
                        message=f"Report generated successfully"
                    )
                else:
                    report.mark_failed(file_result.message)
                    return file_result
                    
            except Exception as e:
                report.mark_failed(str(e))
                raise
                
        except ReportTemplate.DoesNotExist:
            return ServiceResult.error("Report template not found")
        except Exception as e:
            return ServiceResult.error(f"Failed to generate report: {str(e)}")
    
    def _get_report_data(self, template: ReportTemplate, 
                        filters: Dict[str, Any]) -> ServiceResult:
        """Get data for report based on template type"""
        try:
            if template.report_type == 'STOCK_SUMMARY':
                return self._get_stock_summary_data(filters)
            elif template.report_type == 'STOCK_VALUATION':
                return self._get_stock_valuation_data(filters)
            elif template.report_type == 'MOVEMENT_HISTORY':
                return self._get_movement_history_data(filters)
            elif template.report_type == 'ABC_ANALYSIS':
                return self._get_abc_analysis_data(filters)
            elif template.report_type == 'AGING_ANALYSIS':
                return self._get_aging_analysis_data(filters)
            elif template.report_type == 'REORDER_REPORT':
                return self._get_reorder_report_data(filters)
            elif template.report_type == 'DEAD_STOCK':
                return self._get_dead_stock_data(filters)
            elif template.report_type == 'FAST_SLOW_MOVING':
                return self._get_fast_slow_moving_data(filters)
            elif template.report_type == 'SUPPLIER_PERFORMANCE':
                return self._get_supplier_performance_data(filters)
            elif template.report_type == 'PURCHASE_ANALYSIS':
                return self._get_purchase_analysis_data(filters)
            elif template.report_type == 'CUSTOM':
                return self._get_custom_report_data(template, filters)
            else:
                return ServiceResult.error(f"Unsupported report type: {template.report_type}")
                
        except Exception as e:
            (e)}")
    
    def _get_stock_summary_data(self, filters: Dict[str, Any]) -> ServiceResult:
        """Get stock summary report data"""
        try:
            queryset = StockItem.objects.filter(tenant=self.tenant)
            
            # Apply filters
            if filters.get('warehouse_ids'):
                queryset = queryset.filter(warehouse_id__in=filters['warehouse_ids'])
            
            if filters.get('product_ids'):
                queryset = queryset.filter(product_id__in=filters['product_ids'])
            
            if filters.get('category_ids'):
                queryset = queryset.filter(product__category_id__in=filters['category_ids'])
            
            # Get stock data with annotations
            stock_data = queryset.select_related(
                'product', 'warehouse', 'product__category'
            ).annotate(
                available_quantity=F('quantity_on_hand') - F('quantity_reserved'),
                stock_value=F('quantity_on_hand') * F('unit_cost')
            ).values(
                'id', 'product__name', 'product__sku', 'product__category__name',
                'warehouse__name', 'quantity_on_hand', 'quantity_reserved',
                'available_quantity', 'unit_cost', 'stock_value',
                'reorder_level', 'maximum_stock_level', 'abc_classification'
            )
            
            # Calculate totals
            totals = queryset.aggregate(
                total_items=Count('id'),
                total_quantity=Sum('quantity_on_hand'),
                total_reserved=Sum('quantity_reserved'),
                total_value=Sum(F('quantity_on_hand') * F('unit_cost'))
            )
            
            return ServiceResult.success(data={
                'data': list(stock_data),
                'totals': totals,
                'filters_applied': filters
            })
            
        except Exception as e:
            return ServiceResult.error(f"Failed to get stock summary data: {str(e)}")
    
    def _get_stock_valuation_data(self, filters: Dict[str, Any]) -> ServiceResult:
        """Get stock valuation report data"""
        try:
            queryset = StockItem.objects.filter(tenant=self.tenant)
            
            # Apply filters
            if filters.get('warehouse_ids'):
                queryset = queryset.filter(warehouse_id__in=filters['warehouse_ids'])
            
            if filters.get('as_of_date'):
                # For historical valuation, we'd need to calculate based on that date
                pass
            
            # Get valuation data
            valuation_data = queryset.select_related(
                'product', 'warehouse'
            ).annotate(
                current_value=F('quantity_on_hand') * F('unit_cost'),
                available_value=(F('quantity_on_hand') - F('quantity_reserved')) * F('unit_cost')
            ).values(
                'product__name', 'product__sku', 'warehouse__name',
                'quantity_on_hand', 'quantity_reserved', 'unit_cost',
                'current_value', 'available_value'
            )
            
            # Calculate totals by category
            category_totals = queryset.values(
                'product__category__name'
            ).annotate(
                total_quantity=Sum('quantity_on_hand'),
                total_value=Sum(F('quantity_on_hand') * F('unit_cost'))
            ).order_by('-total_value')
            
            # Overall totals
            overall_totals = queryset.aggregate(
                total_items=Count('id'),
                total_quantity=Sum('quantity_on_hand'),
                total_value=Sum(F('quantity_on_hand') * F('unit_cost')),
                total_available_value=Sum(
                    (F('quantity_on_hand') - F('quantity_reserved')) * F('unit_cost')
                )
            )
            
            return ServiceResult.success(data={
                'data': list(valuation_data),
                'category_totals': list(category_totals),
                'overall_totals': overall_totals,
                'filters_applied': filters
            })
            
        except Exception as e: {str(e)}")
    
    def _get_movement_history_data(self, filters: Dict[str, Any]) -> ServiceResult:
        """Get stock movement history data"""
        try:
            start_date = filters.get('start_date', timezone.now() - timedelta(days=30))
            end_date = filters.get('end_date', timezone.now())
            
            queryset = StockMovement.objects.filter(
                tenant=self.tenant,
                created_at__range=[start_date, end_date]
            )
            
            # Apply filters
            if filters.get('warehouse_ids'):
                queryset = queryset.filter(warehouse_id__in=filters['warehouse_ids'])
            
            if filters.get('movement_types'):
                queryset = queryset.filter(movement_type__in=filters['movement_types'])
            
            # Get movement data
            movement_data = queryset.select_related(
                'warehouse', 'created_by'
            ).prefetch_related(
                'items__stock_item__product'
            ).values(
                'id', 'movement_type', 'reference_number', 'warehouse__name',
                'created_at', 'created_by__username', 'notes', 'status'
            )
            
            # Get detailed items for each movement
            detailed_data = []
            for movement in queryset:
                for item in movement.items.all():
                    detailed_data.append({
                        'movement_date': movement.created_at,
                        'movement_type': movement.get_movement_type_display(),
                        'reference_number': movement.reference_number,
                        'product_name': item.stock_item.product.name,
                        'product_sku': item.stock_item.product.sku,
                        'warehouse': movement.warehouse.name,
                        'quantity': item.quantity,
                        'unit_cost': item.unit_cost,
                        'total_value': item.quantity * item.unit_cost,
                        'created_by': movement.created_by.username if movement.created_by else '',
                        'notes': item.notes or movement.notes
                    })
            
            return ServiceResult.success(data={
                'data': detailed_data,
                'summary': list(movement_data),
                'filters_applied': filters
            })
            
        except Exception as e:
            return ServiceResult.error(f"Failed to get movement history: {str(e)}")
    
    def _get_abc_analysis_data(self, filters: Dict[str, Any]) -> ServiceResult:
        """Get ABC analysis data"""
        try:
            queryset = StockItem.objects.filter(tenant=self.tenant)
            
            # Apply filters
            if filters.get('warehouse_ids'):
                queryset = queryset.filter(warehouse_id__in=filters['warehouse_ids'])
            
            # Get ABC analysis data
            abc_data = queryset.select_related(
                'product', 'warehouse'
            ).annotate(
                stock_value=F('quantity_on_hand') * F('unit_cost')
            ).values(
                'product__name', 'product__sku', 'warehouse__name',
                'quantity_on_hand', 'unit_cost', 'stock_value',
                'abc_classification', 'movement_count'
            ).order_by('-stock_value')
            
            # Calculate ABC totals
            abc_totals = queryset.values('abc_classification').annotate(
                item_count=Count('id'),
                total_value=Sum(F('quantity_on_hand') * F('unit_cost')),
                total_quantity=Sum('quantity_on_hand')
            ).order_by('abc_classification')
            
            return ServiceResult.success(data={
                'data': list(abc_data),
                'abc_totals': list(abc_totals),
                'filters_applied': filters
            })
            
        except Exception as e:
            return ServiceResult.error(f"Failed to get ABC analysis: {str(e)}")
    
    def _get_aging_analysis_data(self, filters: Dict[str, Any]) -> ServiceResult:
        """Get stock aging analysis data"""
        try:
            from ...managers.query_utils import InventoryQueryUtils
            
            warehouse_id = filters.get('warehouse_id')
            days_buckets = filters.get('days_buckets', [30, 60, 90, 180, 365])
            
            aging_data = InventoryQueryUtils.get_stock_aging_analysis(
                tenant=self.tenant,
                warehouse=warehouse_id,
                days_buckets=days_buckets
            )
            
            return ServiceResult.success(data={
                'data': list(aging_data),
                'filters_applied': filters
            })
            
        except Exception as e:
            return ServiceResult.error(f"Failed to get aging analysis: {str(e)}")
    
    def _get_reorder_report_data(self, filters: Dict[str, Any]) -> ServiceResult:
        """Get reorder report data"""
        try:
            queryset = StockItem.objects.filter(
                tenant=self.tenant,
                quantity_on_hand__lte=F('reorder_level'),
                reorder_level__gt=0,
                product__is_active=True,
                product__is_purchasable=True
            )
            
            # Apply filters
            if filters.get('warehouse_ids'):
                queryset = queryset.filter(warehouse_id__in=filters['warehouse_ids'])
            
            # Get reorder data with supplier info
            reorder_data = queryset.select_related(
                'product', 'warehouse'
            ).prefetch_related(
                'product__productsupplier_set__supplier'
            ).annotate(
                suggested_quantity=F('maximum_stock_level') - F('quantity_on_hand')
            )
            
            detailed_data = []
            for stock
                primary_supplier = stock_item.product.productsupplier_set.filter(
                    is_primary=True
                ).first()
                
                detailed_data.append({
                    'product_name': stock_item.product.name,
                    'product_sku': stock_item.product.sku,
                    'warehouse': stock_item.warehouse.name,
                    'current_stock': stock_item.quantity_on_hand,
                    'reorder_level': stock_item.reorder_level,
                    'maximum_level': stock_item.maximum_stock_level,
                    'suggested_quantity': max(0, stock_item.maximum_stock_level - stock_item.quantity_on_hand),
                    'primary_supplier': primary_supplier.supplier.name if primary_supplier else 'N/A',
                    'supplier_cost': primary_supplier.supplier_cost if primary_supplier else 0,
                    'estimated_cost': (stock_item.maximum_stock_level - stock_item.quantity_on_hand) * 
                                    (primary_supplier.supplier_cost if primary_supplier else 0),
                    'lead_time_days': primary_supplier.lead_time_days if primary_supplier else 0
                })
            
            return ServiceResult.success(data={
                'data': detailed_data,
                'filters_applied': filters
            })
            
        except Exception as e:
            return ServiceResult.error(f"Failed to get reorder report: {str(e)}")
    
    def _get_dead_stock_data(self, filters: Dict[str, Any]) -> ServiceResult:
        """Get dead stock report data"""
        try:
            days_threshold = filters.get('days_threshold', 180)
            cutoff_date = timezone.now() - timedelta(days=days_threshold)
            
            queryset = StockItem.objects.filter(
                tenant=self.tenant,
                quantity_on_hand__gt=0
            ).filter(
                Q(last_movement_date__lt=cutoff_date) | Q(last_movement_date__isnull=True)
            )
            
            # Apply filters
            if filters.get('warehouse_ids'):
                queryset = queryset.filter(warehouse_id__in=filters['warehouse_ids'])
            
            dead_stock_data = queryset.select_related(
                'product', 'warehouse'
            ).annotate(
                stock_value=F('quantity_on_hand') * F('unit_cost'),
                days_since_movement=Case(
                    When(last_movement_date__isnull=True, 
                         then=Value(999)),
                    default=timezone.now().date() - F('last_movement_date'),
                    output_field=models.IntegerField()
                )
            ).values(
                'product__name', 'product__sku', 'warehouse__name',
                'quantity_on_hand', 'unit_cost', 'stock_value',
                'last_movement_date', 'days_since_movement'
            ).order_by('-stock_value')
            
            return ServiceResult.success(data={
                'data': list(dead_stock_data),
                'filters_applied': filters
            })
            
        except Exception as e:
            return ServiceResult.error(f"Failed to get dead stock data: {str(e)}")
    
    def _get_fast_slow_moving_data(self, filters: Dict[str, Any]) -> ServiceResult:
        """Get fast/slow moving items data"""
        try:
            analysis_days = filters.get('analysis_days', 90)
            cutoff_date = timezone.now() - timedelta(days=analysis_days)
            
            queryset = StockItem.objects.filter(tenant=self.tenant)
            
            # Apply filters
            if filters.get('warehouse_ids'):
                queryset = queryset.filter(warehouse_id__in=filters['warehouse_ids'])
            
            # Annotate with movement metrics
            movement_data = queryset.select_related(
                'product', 'warehouse'
            ).annotate(
                total_movements=Count(
                    'stockmovementitem__movement',
                    filter=Q(stockmovementitem__movement__created_at__gte=cutoff_date)
                ),
                total_quantity_moved=Coalesce(
                    Sum(
                        'stockmovementitem__quantity',
                        filter=Q(stockmovementitem__movement__created_at__gte=cutoff_date)
                    ), 0
                ),
                avg_daily_movement=F('total_quantity_moved') / analysis_days,
                velocity_ratio=Case(
                    When(quantity_on_hand=0, then=Value(0)),
                    default=F('avg_daily_movement') / F('quantity_on_hand'),
                    output_field=models.DecimalField(max_digits=10, decimal_places=4)
                ),
                stock_value=F('quantity_on_hand') * F('unit_cost')
            ).values(
                'product__name', 'product__sku', 'warehouse__name',
                'quantity_on_hand', 'stock_value', 'total_movements',
                'total_quantity_moved', 'avg_daily_movement', 'velocity_ratio'
            ).order_by('-velocity_ratio')
            
            return ServiceResult.success(data={
                'data': list(movement_data),
                'filters_applied': filters
            })
            
        except Exception as e:
            return ServiceResult.error(f"Failed to get fast/slow moving data: {str(e)}")
    
    def _get_supplier_performance_data(self, filters: Dict[str, Any]) -> ServiceResult:
        """Get supplier performance data"""
        try:
            from ...managers.query_utils import InventoryQueryUtils
            
            start_date = filters.get('start_date', timezone.now() - timedelta(days=90))
            end_date = filters.get('end_date', timezone.now())
            
            supplier_data = InventoryQueryUtils.get_supplier_performance_metrics(
                tenant=self.tenant,
                start_date=start_date,
                end_date=end_date
            )
            
            return ServiceResult.success(data={
                'data': list(supplier_data),
                'filters_applied': filters
            })
            
        except Exception as e:
            return ServiceResult.error(f"Failed to get supplier performance: {str(e)}")
    
    def _get_purchase_analysis_data(self, filters: Dict[str, Any]) -> ServiceResult:
        """Get purchase analysis data"""
        try:
            start_date = filters.get('start_date', timezone.now() - timedelta(days=90))
            end_date = filters.get('end_date', timezone.now())
            
            queryset = PurchaseOrder.objects.filter(
                tenant=self.tenant,
                order_date__range=[start_date.date(), end_date.date()]
            )
            
            # Get purchase data
            purchase_data = queryset.select_related(
                'supplier', 'warehouse'
            ).values(
                'po_number', 'supplier__name', 'warehouse__name',
                'order_date', 'expected_delivery_date', 'status',
                'subtotal', 'tax_amount', 'total_amount'
            )
            
            # Get summary by supplier
            supplier_summary = queryset.values(
                'supplier__name'
            ).annotate(
                order_count=Count('id'),
                total_value=Sum('total_amount'),
                avg_order_value=Avg('total_amount')
            ).order_by('-total_value')
            
            # Get monthly trends
            monthly_trends = queryset.extra(
                select={'month': "DATE_TRUNC('month', order_date)"}
            ).values('month').annotate(
                order_count=Count('id'),
                total_value=Sum('total_amount')
            ).order_by('month')
            
            return ServiceResult.success(data={
                'data': list(purchase_data),
                'supplier_summary': list(supplier_summary),
                'monthly_trends': list(monthly_trends),
                'filters_applied': filters
            })
            
        except Exception as e:
            return ServiceResult.error(f"Failed to get purchase analysis: {str(e)}")
    
    def _generate_report_file(self, report: InventoryReport, data: Dict[str, Any],
                             output_format: str) -> ServiceResult:
        """Generate report file in specified format"""
        try:
            if output_format == 'PDF':
                return self._generate_pdf_report(report, data)
            elif output_format == 'EXCEL':
                return self._generate_excel_report(report, data)
            elif output_format == 'CSV':
                return self._generate_csv_report(report, data)
            elif output_format == 'JSON':
                return self._generate_json_report(report, data)
            elif output_format == 'HTML':
                return self._generate_html_report(report, data)
            else:
                return ServiceResult.error(f"Unsupported output format: {output_format}")
                
        except Exception as e:
            return ServiceResult.error(f"Failed to generate {output_format} file: {str(e)}")
    
    def _generate_pdf_report(self, report: str, Any]) -> ServiceResult:
        """Generate PDF report"""
        if not REPORTLAB_AVAILABLE:
            return ServiceResult.error("PDF generation not available - install reportlab")
        
        try:
            # Create file path
            file_path = f"reports/{report.id}_{report.template.name.lower().replace(' ', '_')}.pdf"
            full_path = f"{settings.MEDIA_ROOT}/{file_path}"
            
            # Ensure directory exists
            import os
            os.makedirs(os.path.dirname(full_path), exist_ok=True)
            
            # Create PDF
            doc = SimpleDocTemplate(full_path, pagesize=A4)
            styles = getSampleStyleSheet()
            story = []
            
            # Title
            title = Paragraph(f"{report.template.name}", styles['Title'])
            story.append(title)
            
            # Report info
            info_text = f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M')}<br/>"
            info_text += f"Generated by: {report.created_by.get_full_name() if report.created_by else 'System'}"
            info = Paragraph(info_text, styles['Normal'])
            story.append(info)
            story.append(Paragraph("<br/><br/>", styles['Normal']))
            
            # Data table
            if data.get('data'):
                table_data = []
                # Headers
                headers = list(data['data'][0].keys()) if data['data'] else []
                table_data.append(headers)
                
                # Data rows
                for row in data['data'][:100]:  # Limit to 100 rows for PDF
                    table_data.append([str(row.get(h, '')) for h in headers])
                
                # Create table
                table = Table(table_data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(table)
            
            # Build PDF
            doc.build(story)
            
            # Get file info
            file_size = os.path.getsize(full_path)
            file_hash = self._calculate_file_hash(full_path)
            
            return ServiceResult.success(data={
                'file_path': file_path,
                'file_size': file_size,
                'file_hash': file_hash
            })
            
        except Exception as e:
            return ServiceResult.error(f"Failed to generate PDF: {str(e)}")
    
    def _generate_[str, Any]) -> ServiceResult:
        """Generate Excel report"""
        if not OPENPYXL_AVAILABLE:
            return ServiceResult.error("Excel generation not available - install openpyxl")
        
        try:
            # Create file path
            file_path = f"reports/{report.id}_{report.template.name.lower().replace(' ', '_')}.xlsx"
            full_path = f"{settings.MEDIA_ROOT}/{file_path}"
            
            # Ensure directory exists
            import os
            os.makedirs(os.path.dirname(full_path), exist_ok=True)
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Report Data"
            
            # Add headers
            if data.get('data'):
                headers = list(data['data'][0].keys()) if data['data'] else []
                
                # Style headers
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                
                # Add data
                for row_idx, row_data in enumerate(data['data'], 2):
                    for col_idx, header in enumerate(headers, 1):
                        ws.cell(row=row_idx, column=col_idx, value=row_data.get(header))
            
            # Save workbook
            wb.save(full_path)
            
            # Get file info
            file_size = os.path.getsize(full_path)
            file_hash = self._calculate_file_hash(full_path)
            
            return ServiceResult.success(data={
                'file_path': file_path,
                'file_size': file_size,
                'file_hash': file_hash
            })
            
        except Exception as e:
            return ServiceResult.error(f"Failed to generate Excel: {str(e)}")
    
    def _generate_csv_report(self, report: InventoryReport, CSV report"""
        try:
            # Create file path
            file_path = f"reports/{report.id}_{report.template.name.lower().replace(' ', '_')}.csv"
            full_path = f"{settings.MEDIA_ROOT}/{file_path}"
            
            # Ensure directory exists
            import os
            os.makedirs(os.path.dirname(full_path), exist_ok=True)
            
            # Write CSV
            with open(full_path, 'w', newline='', encoding='utf-8') as csvfile:
                if data.get('data'):
                    fieldnames = list(data['data'][0].keys()) if data['data'] else []
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    
                    writer.writeheader()
                    for row in data['data']:
                        writer.writerow(row)
            
            # Get file info
            file_size = os.path.getsize(full_path)
            file_hash = self._calculate_file_hash(full_path)
            
            return ServiceResult.success(data={
                'file_path': file_path,
                'file_size': file_size,
                'file_hash': file_hash
            })
            
        except Exception as e:
            return ServiceResult.error(f"Failed to generate CSV: {str(e)}")
    
    def _generate_json_report(self, report: InventoryReport, data: Dict[str, Any]) -> ServiceResult:
        """Generate JSON report"""
        try:
            # Create file path
            file_path = f"reports/{report.id}_{report.template.name.lower().replace(' ', '_')}.json"
            full_path = f"{settings.MEDIA_ROOT}/{file_path}"
            
            # Ensure directory exists
            import os
            os.makedirs(os.path.dirname(full_path), exist_ok=True)
            
            # Prepare JSON data
            json_data = {
                'report_info': {
                    'name': report.template.name,
                    'generated_at': timezone.now().isoformat(),
                    'generated_by': report.created_by.username if report.created_by else 'System',
                    'filters_applied': report.filters_applied
                },
                'data': data['data'],
                'totals': data.get('totals', {}),
                'summary': data.get('summary', {})
            }
            
            # Write JSON
            with open(full_path, 'w', encoding='utf-8') as jsonfile:
                json.dump(json_data, jsonfile, indent=2, default=str)
            
            # Get file info
            file_size = os.path.getsize(full_path)
            file_hash = self._calculate_file_hash(full_path)
            
            return ServiceResult.success(data={
                'file_path': file_path,
                'file_size': file_size,
                'file_hash': file_hash
            })
            
        except Exception as e:
            return ServiceResult.error(f"Failed to generate JSON: {str(e)}")
    
    def _calculate_file_hash(self, file_path: str) -> str:
        """Calculate SHA256 hash of file"""
        hash_sha256 = hashlib.sha256()
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_sha256.update(chunk)
        return hash_sha256.hexdigest()
    
    def get_report_status(self, report_id: int) -> ServiceResult:
        """Get status of a report generation"""
        try:
            report = InventoryReport.objects.get(id=report_id, tenant=self.tenant)
            
            return ServiceResult.success(data={
                'id': report.id,
                'status': report.status,
                'progress': self._calculate_progress(report),
                'file_ready': report.status == 'COMPLETED',
                'file_path': report.file_path if report.status == 'COMPLETED' else None,
                'error_message': report.error_message if report.status == 'FAILED' else None
            })
            
        except InventoryReport.DoesNotExist:
            return ServiceResult.error("Report not found")
        except Exception as e:
            return ServiceResult.error(f"Failed to get report status: {str(e)}")
    
    def _calculate_progress(self, report: InventoryReport) -> int:
        """Calculate report generation progress"""
        if report.status == 'PENDING':
            return 0
        elif report.status == 'GENERATING':
            # Estimate based on time elapsed
            if report.generation_started_at:
                elapsed = (timezone.now() - report.generation_started_at).total_seconds()
                # Assume 60 seconds for completion, cap at 90%
                return min(90, int((elapsed / 60) * 100))
            return 10
        elif report.status == 'COMPLETED':
            return 100
        elif report.status == 'FAILED':
            return 0
        else:
            return 0