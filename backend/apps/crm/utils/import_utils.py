# crm/utils/import_utils.py
"""
Data Import Utilities for CRM Module

Provides comprehensive data import capabilities including:
- Multi-format import (CSV, Excel, JSON)
- Data validation and cleaning
- Duplicate detection and handling
- Field mapping and transformation
- Bulk import with progress tracking
- Error handling and reporting
- Import templates and validation
"""

import csv
import json
import io
import re
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from typing import Dict, Any, List, Optional, Union, Tuple, Generator
from dataclasses import dataclass, field

import openpyxl
import pandas as pd
from django.db import transaction, IntegrityError
from django.core.exceptions import ValidationError
from django.utils import timezone
from django.contrib.auth import get_user_model

from crm.utils.validators import (
    validate_email_format, validate_phone_number, validate_currency_amount
)
from crm.utils.formatters import format_phone_number, parse_name_components

User = get_user_model()


@dataclass
class ImportConfiguration:
    """Configuration for data import operations."""
    source_format: str  # csv, excel, json
    target_model: str  # Model name to import to
    field_mapping: Dict[str, str] = field(default_factory=dict)
    required_fields: List[str] = field(default_factory=list)
    duplicate_handling: str = 'skip'  # skip, update, create
    duplicate_check_fields: List[str] = field(default_factory=list)
    validation_rules: Dict[str, Any] = field(default_factory=dict)
    data_transformations: Dict[str, str] = field(default_factory=dict)
    batch_size: int = 100
    skip_errors: bool = True
    create_missing_references: bool = False


@dataclass
class ImportResult:
    """Result of data import operation."""
    total_records: int = 0
    successful_imports: int = 0
    failed_imports: int = 0
    skipped_records: int = 0
    updated_records: int = 0
    errors: List[Dict[str, Any]] = field(default_factory=list)
    warnings: List[Dict[str, Any]] = field(default_factory=list)
    import_id: str = ""
    processing_time: float = 0.0


class CRMDataImporter:
    """
    Main class for handling CRM data imports.
    """
    
    def __init__(self, tenant=None):
        self.tenant = tenant
        self.supported_formats = ['csv', 'excel', 'json']
        self.model_mapping = {
            'lead': 'crm.models.lead.Lead',
            'opportunity': 'crm.models.opportunity.Opportunity',
            'account': 'crm.models.account.Account',
            'contact': 'crm.models.account.Contact',
            'activity': 'crm.models.activity.Activity'
        }
    
    def import_data(self,
                   data_source: Union[str, io.IOBase],
                   config: ImportConfiguration) -> ImportResult:
        """
        Import data from various sources.
        
        Args:
            data_source: File path, file object, or data string
            config: Import configuration
        
        Returns:
            ImportResult: Results of the import operation
        """
        import_start_time = timezone.now()
        import_id = f"import_{import_start_time.strftime('%Y%m%d_%H%M%S')}"
        
        result = ImportResult(import_id=import_id)
        
        try:
            # Parse data based on format
            if config.source_format == 'csv':
                raw_data = self._parse_csv_data(data_source)
            elif config.source_format == 'excel':
                raw_data = self._parse_excel_data(data_source)
            elif config.source_format == 'json':
                raw_data = self._parse_json_data(data_source)
            else:
                raise ValueError(f"Unsupported format: {config.source_format}")
            
            result.total_records = len(raw_data)
            
            # Process data in batches
            processed_data = []
            for i in range(0, len(raw_data), config.batch_size):
                batch = raw_data[i:i + config.batch_size]
                batch_result = self._process_batch(batch, config, i)
                
                result.successful_imports += batch_result.successful_imports
                result.failed_imports += batch_result.failed_imports
                result.skipped_records += batch_result.skipped_records
                result.updated_records += batch_result.updated_records
                result.errors.extend(batch_result.errors)
                result.warnings.extend(batch_result.warnings)
            
            # Calculate processing time
            result.processing_time = (timezone.now() - import_start_time).total_seconds()
            
            # Log import results
            self._log_import_results(result, config)
            
        except Exception as e:
            result.errors.append({
                'row': 0,
                'field': 'general',
                'error': f"Import failed: {str(e)}"
            })
        
        return result
    
    def _parse_csv_data(self, data_source: Union[str, io.IOBase]) -> List[Dict[str, Any]]:
        """Parse CSV data into list of dictionaries."""
        if isinstance(data_source, str):
            # File path
            with open(data_source, 'r', encoding='utf-8') as file:
                return self._read_csv_file(file)
        else:
            # File object or string
            if hasattr(data_source, 'read'):
                return self._read_csv_file(data_source)
            else:
                # String data
                return self._read_csv_file(io.StringIO(data_source))
    
    def _read_csv_file(self, file_obj) -> List[Dict[str, Any]]:
        """Read CSV file object."""
        csv_reader = csv.DictReader(file_obj)
        return [row for row in csv_reader]
    
    def _parse_excel_data(self, data_source: Union[str, io.IOBase]) -> List[Dict[str, Any]]:
        """Parse Excel data into list of dictionaries."""
        if isinstance(data_source, str):
            # File path
            workbook = openpyxl.load_workbook(data_source)
        else:
            # File object
            workbook = openpyxl.load_workbook(data_source)
        
        # Use first worksheet
        worksheet = workbook.active
        
        # Get headers from first row
        headers = []
        for cell in worksheet[1]:
            headers.append(cell.value)
        
        # Read data rows
        data = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            row_data = {}
            for i, value in enumerate(row):
                if i < len(headers) and headers[i]:
                    row_data[headers[i]] = value
            if any(row_data.values()):  # Skip empty rows
                data.append(row_data)
        
        return data
    
    def _parse_json_data(self, data_source: Union[str, io.IOBase]) -> List[Dict[str, Any]]:
        """Parse JSON data into list of dictionaries."""
        if isinstance(data_source, str):
            if data_source.endswith('.json'):
                # File path
                with open(data_source, 'r', encoding='utf-8') as file:
                    json_data = json.load(file)
            else:
                # JSON string
                json_data = json.loads(data_source)
        else:
            # File object
            json_data = json.load(data_source)
        
        # Handle different JSON structures
        if isinstance(json_data, list):
            return json_data
        elif isinstance(json_data, dict):
            return json_data['data']
            else:
                return [json_data]
        else:
            raise ValueError("Invalid JSON structure")
    
    def _process_batch(self, 
                      batch_ Any]], 
                      config: ImportConfiguration,
                      batch_offset: int) -> ImportResult:
        """Process a batch of import data."""
        batch_result = ImportResult()
        
        with transaction.atomic():
            for row_index, row_data in enumerate(batch_data):
                absolute_row_index = batch_offset + row_index + 1
                
                try:
                    # Transform and validate data
                    processed_row = self._process_row(row_data, config, absolute_row_index)
                    
                    if processed_row is None:
                        batch_result.skipped_records += 1
                        continue
                    
                    # Check for duplicates
                    duplicate_action = self._check_duplicates(processed_row, config)
                    
                    if duplicate_action == 'skip':
                        batch_result.skipped_records += 1
                        batch_result.warnings.append({
                            'row': absolute_row_index,
                            'message': 'Record skipped due to duplicate'
                        })
                        continue
                    elif duplicate_action == 'update':
                        self._update_existing_record(processed_row, config)
                        batch_result.updated_records += 1
                    else:  # create
                        self._create_new_record(processed_row, config)
                        batch_result.successful_imports += 1
                
                except Exception as e:
                    batch_result.failed_imports += 1
                    batch_result.errors.append({
                        'row': absolute_row_index,
                        'data': row_data,
                        'error': str(e)
                    })
                    
                    if not config.skip_errors:
                        raise
        
        return batch_result
    
    def _process_row(self, [str, Any], 
                    config: ImportConfiguration,
                    row_index: int) -> Optional[Dict[str, Any]]:
        """Process and validate a single row of data."""
        processed_row = {}
        
        # Apply field mapping
        for source_field, target_field in config.field_mapping.items():
            if source_field in row_ = row_data[source_field]
                
                # Apply data transformations
                if target_field in config.data_transformations:
                    value = self._apply_transformation(
                        value, 
                        config.data_transformations[target_field]
                    )
                
                processed_row[target_field] = value
        
        # Validate required fields
        for required_field in config.required_fields:
            if not processed_row.get(required_field):
                raise ValidationError(f"Required field '{required_field}' is missing or empty")
        
        # Apply validation rules
        self._validate_row_data(processed_row, config.validation_rules, row_index)
        
        # Add tenant information
        if self.tenant:
            processed_row['tenant'] = self.tenant
        
        return processed_row
    
    def _apply_transformation(self, value: Any, transformation: str) -> Any:
        """Apply data transformation to field value."""
        if value is None or value == '':
            return None
        
        transformations = {
            'uppercase': lambda x: str(x).upper(),
            'lowercase': lambda x: str(x).lower(),
            'title_case': lambda x: str(x).title(),
            'strip_whitespace': lambda x: str(x).strip(),
            'parse_phone': lambda x: format_phone_number(str(x)),
            'parse_email': lambda x: str(x).lower().strip(),
            'parse_date': self._parse_date_value,
            'parse_decimal': lambda x: Decimal(str(x)) if x else None,
            'parse_boolean': self._parse_boolean_value,
            'clean_text': lambda x: re.sub(r'[^\w\s-]', '', str(x)).strip()
        }
        
        if transformation in transformations:
            try:
                return transformations[transformation](value)
            except Exception as e:
                raise ValidationError(f"Transformation '{transformation}' failed: {e}")
        
        return value
    
    def _parse_date_value(self, value: Any) -> Optional[date]:
        """Parse various date formats."""
        if not value:
            return None
        
        if isinstance(value, (date, datetime)):
            return value.date() if isinstance(value, datetime) else value
        
        # Try common date formats
        date_formats = [
            '%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S',
            '%m-%d-%Y', '%d-%m-%Y', '%Y/%m/%d', '%d.%m.%Y'
        ]
        
        date_str = str(value).strip()
        for date_format in date_formats:
            try:
                return datetime.strptime(date_str, date_format).date()
            except ValueError:
                continue
        
        raise ValidationError(f"Unable to parse date: {value}")
    
    def _parse_boolean_value(self, value: Any) -> bool:
        """Parse boolean values from various formats."""
        if isinstance(value, bool):
            return value
        
        if isinstance(value, (int, float)):
            return bool(value)
        
        if isinstance(value, str):
            value = value.lower().strip()
            if value in ['true', 'yes', '1', 'on', 'y']:
                return True
            elif value in ['false', 'no', '0', 'off', 'n', '']:
                return False
        
        raise ValidationError(f"Unable to parse boolean: {value}")
    
    def _validate_row_data(self, [str, Any], 
                          validation_rules: Dict[str, Any],
                          row_index: int):
        """Validate row data against validation rules."""
        for field_name, value in row_data.items():
            if field_name in validation_rules:
                rule = validation_rules[field_name]
                
                try:
                    if rule == 'email':
                        if value:
                            validate_email_format(value)
                    elif rule == 'phone':
                        if value:
                            validate_phone_number(value)
                    elif rule == 'currency':
                        if value:
                            validate_currency_amount(value)
                    elif isinstance(rule, dict):
                        # Custom validation rule
                        if 'min_length' in rule and len(str(value or '')) < rule['min_length']:
                            raise ValidationError(f"Field '{field_name}' is too short")
                        if 'max_length' in rule and len(str(value or '')) > rule['max_length']:
                            raise ValidationError(f"Field '{field_name}' is too long")
                        if 'regex' in rule and value and not re.match(rule['regex'], str(value)):
                            raise ValidationError(f"Field '{field_name}' doesn't match required pattern")
                
                except ValidationError as e:
                    raise ValidationError(f"Row {row_index}, Field '{field_name}': {e}")
    
    def _check_duplicates(self, 
                         row_ 
                         config: ImportConfiguration) -> str:
        """Check for duplicate records and return action to take."""
        if not config.duplicate_check_fields:
            return 'create'
        
        # Get target model
        model_class = self._get_model_class(config.target_model)
        
        # Build duplicate check query
        filters = {}
        for field_name in config.duplicate_check_fields:
            if field_name in row_data and row_data[field_name]:
                filters[field_name] = row_data[field_name]
        
        if not filters:
            return 'create'
        
        # Add tenant filter if applicable
        if self.tenant and hasattr(model_class, 'tenant'):
            filters['tenant'] = self.tenant
        
        # Check for existing record
        existing_record = model_class.objects.filter(**filters).first()
        
        if existing_record:
            if config.duplicate_handling == 'skip':
                return 'skip'
            elif config.duplicate_handling == 'update':
                # Store existing record for update
                row_data['_existing_record'] = existing_record
                return 'update'
            else:  # create duplicate
                return 'create'
        
        return 'create'
    
    def _create_new_record(self, row], config: ImportConfiguration):
        """Create new record in database."""
        model_class = self._get_model_class(config.target_model)
        
        # Remove internal fields
        create_data = {k: v for k, v in row_data.items() if not k.startswith('_')}
        
        # Handle foreign key relationships
        create_data = self._resolve_foreign_keys(create_data, model_class, config)
        
        # Create record
        model_class.objects.create(**create_data)
    
    def _update_existing_record(self, row], config: ImportConfiguration):
        """Update existing record in database."""
        existing_record = row_data.pop('_existing_record')
        
        # Remove internal fields
        update_data = {k: v for k, v in row_data.items() if not k.startswith('_')}
        
        # Handle foreign key relationships
        update_data = self._resolve_foreign_keys(update_data, existing_record.__class__, config)
        
        # Update record
        for field_name, value in update_data.items():
            setattr(existing_record, field_name, value)
        
        existing_record.save()
    
    def _resolve_foreign_keys(self, , Any], 
                             model_class,
                             config: ImportConfiguration) -> Dict[str, Any]:
        """Resolve foreign key relationships in import data."""
        resolved_data = data.copy()
        
        # Get model fields
        for field in model_class._meta.fields:
            if field.name in resolved_data and hasattr(field, 'related_model'):
                # This is a foreign key field
                value = resolved_data[field.name]
                
                if value and not isinstance(value, field.related_model):
                    # Try to resolve foreign key
                    related_obj = self._find_related_object(
                        field.related_model, 
                        value, 
                        config.create_missing_references
                    )
                    resolved_data[field.name] = related_obj
        
        return resolved_data
    
    def _find_related_object(self, related_model, value, create_if_missing: bool):
        """Find related object by various criteria."""
        # Try different lookup strategies
        lookups = []
        
        if hasattr(related_model, 'name'):
            lookups.append(('name', value))
        if hasattr(related_model, 'email'):
            lookups.append(('email', value))
        if hasattr(related_model, 'code'):
            lookups.append(('code', value))
        
        # Try primary key if value is numeric
        try:
            pk_value = int(value)
            lookups.append(('pk', pk_value))
        except (ValueError, TypeError):
            pass
        
        # Add tenant filter if applicable
        base_filters = {}
        if self.tenant and hasattr(related_model, 'tenant'):
            base_filters['tenant'] = self.tenant
        
        # Try each lookup
        for field_name, lookup_value in lookups:
            try:
                filters = base_filters.copy()
                filters[field_name] = lookup_value
                obj = related_model.objects.get(**filters)
                return obj
            except (related_model.DoesNotExist, related_model.MultipleObjectsReturned):
                continue
        
        # Create if allowed and it's a simple name-based model
        if create_if_missing and hasattr(related_model, 'name'):
            create_data = {'name': str(value)}
            if self.tenant and hasattr(related_model, 'tenant'):
                create_data['tenant'] = self.tenant
            
            obj, created = related_model.objects.get_or_create(**create_data)
            return obj
        
        raise ValidationError(f"Could not resolve {related_model.__name__}: {value}")
    
    def _get_model_class(self, model_name: str):
        """Get Django model class from model name."""
        if model_name.lower() in self.model_mapping:
            model_path = self.model_mapping[model_name.lower()]
            module_path, class_name = model_path.rsplit('.', 1)
            
            module = __import__(module_path, fromlist=[class_name])
            return getattr(module, class_name)
        else:
            raise ValueError(f"Unknown model: {model_name}")
    
    def _log_import_results(self, result: ImportResult, config: ImportConfiguration):
        """Log import results for audit purposes."""
        try:
            from crm.models.system import ImportLog
            
            ImportLog.objects.create(
                tenant=self.tenant,
                import_id=result.import_id,
                target_model=config.target_model,
                source_format=config.source_format,
                total_records=result.total_records,
                successful_imports=result.successful_imports,
                failed_imports=result.failed_imports,
                skipped_records=result.skipped_records,
                updated_records=result.updated_records,
                processing_time=result.processing_time,
                error_count=len(result.errors),
                warning_count=len(result.warnings),
                import_date=timezone.now()
            )
        except Exception as e:
            print(f"Failed to log import results: {e}")


class CRMImportTemplates:
    """
    Predefined import templates for common CRM data imports.
    """
    
    @staticmethod
    def get_leads_import_config() -> ImportConfiguration:
        """Get import configuration for leads."""
        return ImportConfiguration(
            source_format='csv',
            target_model='lead',
            field_mapping={
                'First Name': 'first_name',
                'Last Name': 'last_name',
                'Email': 'email',
                'Phone': 'phone',
                'Company': 'company',
                'Source': 'source',
                'Status': 'status',
                'Lead Score': 'score'
            },
            required_fields=['first_name', 'last_name', 'email'],
            duplicate_check_fields=['email'],
            duplicate_handling='skip',
            validation_rules={
                'email': 'email',
                'phone': 'phone',
                'score': {'min_value': 0, 'max_value': 100}
            },
            data_transformations={
                'first_name': 'title_case',
                'last_name': 'title_case',
                'email': 'parse_email',
                'phone': 'parse_phone'
            }
        )
    
    @staticmethod
    def get_contacts_import_config() -> ImportConfiguration:
        """Get import configuration for contacts."""
        return ImportConfiguration(
            source_format='csv',
            target_model='contact',
            field_mapping={
                'First Name': 'first_name',
                'Last Name': 'last_name',
                'Email': 'email',
                'Phone': 'phone',
                'Mobile Phone': 'mobile_phone',
                'Job Title': 'title',
                'Company': 'account',
                'Department': 'department'
            },
            required_fields=['first_name', 'last_name'],
            duplicate_check_fields=['email', 'first_name', 'last_name'],
            duplicate_handling='update',
            validation_rules={
                'email': 'email',
                'phone': 'phone',
                'mobile_phone': 'phone'
            },
            data_transformations={
                'first_name': 'title_case',
                'last_name': 'title_case',
                'email': 'parse_email',
                'phone': 'parse_phone',
                'mobile_phone': 'parse_phone'
            },
            create_missing_references=True
        )
    
    @staticmethod
    def get_accounts_import_config() -> ImportConfiguration:
        """Get import configuration for accounts."""
        return ImportConfiguration(
            source_format='csv',
            target_model='account',
            field_mapping={
                'Company Name': 'name',
                'Industry': 'industry',
                'Website': 'website',
                'Phone': 'phone',
                'Annual Revenue': 'annual_revenue',
                'Employees': 'employees'
            },
            required_fields=['name'],
            duplicate_check_fields=['name'],
            duplicate_handling='skip',
            validation_rules={
                'website': {'regex': r'^https?://'},
                'phone': 'phone',
                'annual_revenue': 'currency'
            },
            data_transformations={
                'name': 'title_case',
                'phone': 'parse_phone',
                'annual_revenue': 'parse_decimal'
            }
        )


# Convenience functions
def import_leads_from_csv(file_path: str, tenant=None) -> ImportResult:
    """Import leads from CSV file."""
    importer = CRMDataImporter(tenant)
    config = CRMImportTemplates.get_leads_import_config()
    return importer.import_data(file_path, config)


def import_contacts_from_excel(file_obj: io.IOBase, tenant=None) -> ImportResult:
    """Import contacts from Excel file."""
    importer = CRMDataImporter(tenant)
    config = CRMImportTemplates.get_contacts_import_config()
    config.source_format = 'excel'
    return importer.import_data(file_obj, config)


def validate_import_data(data_source: Union[str, io.IOBase], 
                        config: ImportConfiguration) -> Dict[str, Any]:
    """Validate import data without actually importing."""
    importer = CRMDataImporter()
    
    # Parse data
    if config.source_format == 'csv':
        raw_data = importer._parse_csv_data(data_source)
    elif config.source_format == 'excel':
        raw_data = importer._parse_excel_data(data_source)
    elif config.source_format == 'json':
        raw_data = importer._parse_json_data(data_source)
    else:
        return {'valid': False, 'errors': ['Unsupported format']}
    
    validation_result = {
        'valid': True,
        'total_records': len(raw_data),
        'errors': [],
        'warnings': [],
        'sample_data': raw_data[:5] if raw_data else []
    }
    
    # Validate each row
    for i, row_data in enumerate(raw_data[:100]):  # Validate first 100 rows
        try:
            processed_row = importer._process_row(row_data, config, i + 1)
        except Exception as e:
            validation_result['valid'] = False
            validation_result['errors'].append({
                'row': i + 1,
                'error': str(e)
            })
    
    return validation_result