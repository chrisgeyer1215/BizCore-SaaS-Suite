# crm/utils/export_utils.py
"""
Data Export Utilities for CRM Module

Provides comprehensive data export capabilities including:
- Multi-format export (CSV, Excel, JSON, PDF)
- Bulk data export with filtering
- Custom field mapping and transformation
- Large dataset handling with pagination
- Export templates and scheduling
- Security and permission handling
"""

import csv
import json
import io
import zipfile
from datetime import datetime, date
from decimal import Decimal
from typing import Dict, Any, List, Optional, Union, Generator
from dataclasses import dataclass, asdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

from django.http import HttpResponse, StreamingHttpResponse
from django.utils import timezone
from django.core.serializers import serialize
from django.core.paginator import Paginator
from django.db.models import QuerySet
from django.template.loader import render_to_string


@dataclass
class ExportConfiguration:
    """Configuration for data export operations."""
    format: str  # csv, excel, json, pdf
    include_headers: bool = True
    include__size: int = 1000
    max_records: Optional[int] = None
    field_mapping: Optional[Dict[str, str]] = None
    custom_fields: Optional[List[str]] = None
    filters: Optional[Dict[str, Any]] = None
    sort_fields: Optional[List[str]] = None
    template_name: Optional[str] = None


class CRMDataExporter:
    """
    Main class for handling CRM data exports.
    """
    
    def __init__(self, tenant=None):
        self.tenant = tenant
        self.supported_formats = ['csv', 'excel', 'json', 'pdf']
    
    def export_queryset(self, 
                       queryset: QuerySet,
                       config: ExportConfiguration,
                       filename_prefix: str = "crm_export") -> HttpResponse:
        """
        Export Django queryset to specified format.
        
        Args:
            queryset: Django queryset to export
            config: Export configuration
            filename_prefix: Prefix for export filename
        
        Returns:
            HttpResponse: Response with exported data
        """
        if config.format not in self.supported_formats:
            raise ValueError(f"Unsupported export format: {config.format}")
        
        # Apply filters if provided
        if config.filters:
            queryset = self._apply_filters(queryset, config.filters)
        
        # Apply sorting if provided
        if config.sort_fields:
            queryset = queryset.order_by(*config.sort_fields)
        
        # Limit records if specified
        if config.max_records:
            queryset = queryset[:config.max_records]
        
        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{filename_prefix}_{timestamp}.{config.format}"
        
        # Export based on format
        if config.format == 'csv':
            return self._export_to_csv(queryset, config, filename)
        elif config.format == 'excel':
            return self._export_to_excel(queryset, config, filename)
        elif config.format == 'json':
            return self._export_to_json(queryset, config, filename)
        elif config.format == 'pdf':
            return self._export_to_pdf(queryset, config, filename)
    
    def export_bulk_data(self,
                        data_sources: Dict[str, QuerySet],
                        config: ExportConfiguration,
                        filename: str = "crm_bulk_export") -> HttpResponse:
        """
        Export multiple data sources to a single archive.
        
        Args:
            data_sources: Dictionary of data source name to queryset
            config: Export configuration
            filename: Base filename for archive
        
        Returns:
            HttpResponse: Response with ZIP archive containing all exports
        """
        # Create ZIP archive in memory
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for source_name, queryset in data_sources.items():
                # Export each data source
                if config.format == 'csv':
                    data = self._queryset_to_csv_string(queryset, config)
                    zip_file.writestr(f"{source_name}.csv", data)
                elif config.format == 'excel':
                    data = self._queryset_to_excel_bytes(queryset, config)
                    zip_file.writestr(f"{source_name}.xlsx", data)
                elif config.format == 'json':
                    data = self._queryset_to_json_string(queryset, config)
                    zip_file.writestr(f"{source_name}.json", data)
            
            # Add metadata file
            metadata = {
                'export_date': timezone.now().isoformat(),
                'tenant_id': self.tenant.id if self.tenant else None,
                'data_sources': list(data_sources.keys()),
                'configuration': asdict(config)
            }
            zip_file.writestr('export_metadata.json', json.dumps(metadata, indent=2))
        
        # Prepare response
        zip_buffer.seek(0)
        response = HttpResponse(
            zip_buffer.getvalue(),
            content_type='application/zip'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.zip"'
        return response
    
    def _apply_filters(self, queryset: QuerySet, filters: Dict[str, Any]) -> QuerySet:
        """Apply filters to queryset."""
        filtered_qs = queryset
        
        for field_name, filter_value in filters.items():
            if isinstance(filter_value, dict):
                # Handle complex filters (e.g., date ranges, lookups)
                for lookup, value in filter_value.items():
                    filter_key = f"{field_name}__{lookup}"
                    filtered_qs = filtered_qs.filter(**{filter_key: value})
            else:
                # Simple equality filter
                filtered_qs = filtered_qs.filter(**{field_name: filter_value})
        
        return filtered_qs
    
    def _export_to_csv(self, queryset: QuerySet, config: ExportConfiguration, filename: str) -> HttpResponse:
        """Export queryset to CSV format."""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Use generator for large datasets
        def generate_csv_rows():
            # Get field names
            if queryset.exists():
                model_fields = self._get_model_fields(queryset.model)
                field_names = self._apply_field_mapping(model_fields, config.field_mapping)
                
                # Write header
                if config.include_headers:
                    yield self._format_csv_row(field_names)
                
                # Write data rows in batches
                paginator = Paginator(queryset, config.batch_size)
                for page_num in range(1, paginator.num_pages + 1):
                    page = paginator.get_page(page_num)
                    for obj in page.object_list:
                        row_data = self._extract_object_data(obj, model_fields, config)
                        yield self._format_csv_row(row_data)
        
        response.content = ''.join(generate_csv_rows())
        return response
    
    def _export_to_excel(self, queryset: QuerySet, config: ExportConfiguration, filename: str) -> HttpResponse:
        """Export queryset to Excel format."""
        # Create workbook
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "CRM Export"
        
        if queryset.exists():
            model_fields = self._get_model_fields(queryset.model)
            field_names = self._apply_field_mapping(model_fields, config.field_mapping)
            
            # Write headers with formatting
            if config.include_headers:
                for col_num, header in enumerate(field_names, 1):
                    cell = worksheet.cell(row=1, column=col_num, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
            
            # Write data rows
            row_num = 2 if config.include_headers else 1
            paginator = Paginator(queryset, config.batch_size)
            
            for page_num in range(1, paginator.num_pages + 1):
                page = paginator.get_page(page_num)
                for obj in page.object_list:
                    row_data = self._extract_object_data(obj, model_fields, config)
                    for col_num, value in enumerate(row_data, 1):
                        worksheet.cell(row=row_num, column=col_num, value=value)
                    row_num += 1
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add metadata sheet if requested
        if config.include_metadata:
            self._add_metadata_sheet(workbook, queryset, config)
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Prepare response
        response = HttpResponse(
            excel_buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    def _export_to_json(self, queryset: QuerySet, config: ExportConfiguration, filename: str) -> HttpResponse:
        """Export queryset to JSON format."""
        def json_serializer(obj):
            """JSON serializer for non-serializable objects."""
            if isinstance(obj, (datetime, date)):
                return obj.isoformat()
            elif isinstance(obj, Decimal):
                return float(obj)
            elif hasattr(obj, '__dict__'):
                return obj.__dict__
            else:
                return str(obj)
        
        # Prepare data structure
        export_data = {
            'metadata': {
                'export_date': timezone.now().isoformat(),
                'tenant_id': self.tenant.id if self.tenant else None,
                'record_count': queryset.count(),
                'model': queryset.model.__name__
            } if config.include_metadata else {},
            'data': []
        }
        
        if queryset.exists():
            model_fields = self._get_model_fields(queryset.model)
            
            # Process data in batches
            paginator = Paginator(queryset, config.batch_size)
            for page_num in range(1, paginator.num_pages + 1):
                page = paginator.get_page(page_num)
                for obj in page.object_list:
                    obj_data = {}
                    for field in model_fields:
                        try:
                            value = getattr(obj, field)
                            # Handle foreign keys
                            if hasattr(value, 'pk'):
                                obj_data[field] = {
                                    'id': value.pk,
                                    'name': str(value)
                                }
                            else:
                                obj_data[field] = value
                        except AttributeError:
                            obj_data[field] = None
                    
                    export_data['data'].append(obj_data)
        
        # Prepare response
        json_string = json.dumps(export_data, default=json_serializer, indent=2)
        response = HttpResponse(json_string, content_type='application/json')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    def _export_to_pdf(self, queryset: QuerySet, config: ExportConfiguration, filename: str) -> HttpResponse:
        """Export queryset to PDF format."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        # Add title
        title = Paragraph("CRM Data Export", title_style)
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        if config.include_metadata:
            # Add metadata
            metadata_text = f"""
            <b>Export Date:</b> {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>
            <b>Records:</b> {queryset.count()}<br/>
            <b>Model:</b> {queryset.model.__name__}
            """
            metadata_para = Paragraph(metadata_text, styles['Normal'])
            elements.append(metadata_para)
            elements.append(Spacer(1, 20))
        
        if queryset.exists():
            model_fields = self._get_model_fields(queryset.model)
            field_names = self._apply_field_mapping(model_fields, config.field_mapping)
            
            # Prepare table data
            table_data = []
            
            # Add headers
            if config.include_headers:
                table_data.append(field_names)
            
            # Add data rows (limit for PDF performance)
            max_pdf_records = min(config.max_records or 1000, 1000)
            for obj in queryset[:max_pdf_records]:
                row_data = self._extract_object_data(obj, model_fields, config)
                # Truncate long text for PDF
                row_data = [str(item)[:50] + '...' if len(str(item)) > 50 else str(item) 
                           for item in row_data]
                table_data.append(row_data)
            
            # Create table
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(table)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        # Prepare response
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    def _get_model_fields(self, model) -> List[str]:
        """Get list of model field names."""
        return [field.name for field in model._meta.fields]
    
    def _apply_field_mapping(self, fields: List[str], field_mapping: Optional[Dict[str, str]]) -> List[str]:
        """Apply field mapping to field names."""
        if not field_mapping:
            return fields
        
        return [field_mapping.get(field, field) for field in fields]
    
    def _extract_object_data(self, obj, fields: List[str], config: ExportConfiguration) -> List[Any]:
        """Extract data from object based on field configuration."""
        data = []
        
        for field in fields:
            try:
                value = getattr(obj, field)
                
                # Handle different data types
                if isinstance(value, datetime):
                    data.append(value.strftime('%Y-%m-%d %H:%M:%S'))
                elif isinstance(value, date):
                    data.append(value.strftime('%Y-%m-%d'))
                elif isinstance(value, Decimal):
                    data.append(float(value))
                elif value is None:
                    data.append('')
                elif hasattr(value, 'pk'):  # Foreign key
                    data.append(str(value))
                else:
                    data.append(str(value))
            except AttributeError:
                data.append('')
        
        return data
    
    def _format_csv_row(self, row str:
        """Format row data as CSV string."""
        output = io.StringIO()
        writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)
        writer.writerow(row_data)
        return output.getvalue()
    
    def _queryset_to_csv_string(self, queryset: QuerySet, config: ExportConfiguration) -> str:
        """Convert queryset to CSV string."""
        output = io.StringIO()
        
        if queryset.exists():
            model_fields = self._get_model_fields(queryset.model)
            field_names = self._apply_field_mapping(model_fields, config.field_mapping)
            
            writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)
            
            # Write header
            if config.include_headers:
                writer.writerow(field_names)
            
            # Write data
            for obj in queryset:
                row_data = self._extract_object_data(obj, model_fields, config)
                writer.writerow(row_data)
        
        return output.getvalue()
    
    def _queryset_to_excel_bytes(self, queryset: QuerySet, config: ExportConfiguration) -> bytes:
        """Convert queryset to Excel bytes."""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        
        if queryset.exists():
            model_fields = self._get_model_fields(queryset.model)
            field_names = self._apply_field_mapping(model_fields, config.field_mapping)
            
            # Write headers
            if config.include_headers:
                for col_num, header in enumerate(field_names, 1):
                    worksheet.cell(row=1, column=col_num, value=header)
            
            # Write data
            row_num = 2 if config.include_headers else 1
            for obj in queryset:
                row_data = self._extract_object_data(obj, model_fields, config)
                for col_num, value in enumerate(row_data, 1):
                    worksheet.cell(row=row_num, column=col_num, value=value)
                row_num += 1
        
        # Save to bytes
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def _queryset_to_json_string(self, queryset: QuerySet, config: ExportConfiguration) -> str:
        """Convert queryset to JSON string."""
        data = []
        
        if queryset.exists():
            model_fields = self._get_model_fields(queryset.model)
            
            for obj in queryset:
                obj_data = {}
                for field in model_fields:
                    try:
                        value = getattr(obj, field)
                        if isinstance(value, (datetime, date)):
                            obj_data[field] = value.isoformat()
                        elif isinstance(value, Decimal):
                            obj_data[field] = float(value)
                        elif hasattr(value, 'pk'):
                            obj_data[field] = str(value)
                        else:
                            obj_data[field] = value
                    except AttributeError:
                        obj_data[field] = None
                
                data.append(obj_data)
        
        return json.dumps(data, indent=2)
    
    def _add_metadata_sheet(self, workbook, queryset: QuerySet, config: ExportConfiguration):
        """Add metadata sheet to Excel workbook."""
        metadata_sheet = workbook.create_sheet("Metadata")
        
        metadata = [
            ["Export Date", timezone.now().strftime('%Y-%m-%d %H:%M:%S')],
            ["Tenant ID", self.tenant.id if self.tenant else "N/A"],
            ["Model", queryset.model.__name__],
            ["Record Count", queryset.count()],
            ["Export Format", config.format],
            ["Batch Size", config.batch_size]
        ]
        
        for row_num, (key, value) in enumerate(metadata, 1):
            metadata_sheet.cell(row=row_num, column=1, value=key).font = Font(bold=True)
            metadata_sheet.cell(row=row_num, column=2, value=str(value))
        
        # Auto-adjust column widths
        for column in metadata_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            metadata_sheet.column_dimensions[column_letter].width = adjusted_width


class CRMExportTemplates:
    """
    Predefined export templates for common CRM data exports.
    """
    
    @staticmethod
    def get_leads_export_config() -> ExportConfiguration:
        """Get export configuration for leads."""
        return ExportConfiguration(
            format='excel',
            include_headers=True,
            include_metadata=True,
            field_mapping={
                'first_name': 'First Name',
                'last_name': 'Last Name',
                'email': 'Email Address',
                'phone': 'Phone Number',
                'company': 'Company',
                'source': 'Lead Source',
                'status': 'Status',
                'score': 'Lead Score',
                'assigned_to': 'Assigned To',
                'created_at': 'Created Date',
                'updated_at': 'Last Updated'
            }
        )
    
    @staticmethod
    def get_opportunities_export_config() -> ExportConfiguration:
        """Get export configuration for opportunities."""
        return ExportConfiguration(
            format='excel',
            include_headers=True,
            include_metadata=True,
            field_mapping={
                'name': 'Opportunity Name',
                'account': 'Account',
                'stage': 'Stage',
                'value': 'Value',
                'probability': 'Probability (%)',
                'expected_close_date': 'Expected Close Date',
                'assigned_to': 'Owner',
                'created_at': 'Created Date'
            }
        )
    
    @staticmethod
    def get_contacts_export_config() -> ExportConfiguration:
        """Get export configuration for contacts."""
        return ExportConfiguration(
            format='excel',
            include_headers=True,
            include_metadata=True,
            field_mapping={
                'first_name': 'First Name',
                'last_name': 'Last Name',
                'email': 'Email',
                'phone': 'Phone',
                'mobile_phone': 'Mobile Phone',
                'account': 'Account',
                'title': 'Job Title',
                'department': 'Department',
                'address_street': 'Street Address',
                'address_city': 'City',
                'address_state': 'State',
                'address_postal_code': 'Postal Code',
                'address_country': 'Country'
            }
        )
    
    @staticmethod
    def get_accounts_export_config() -> ExportConfiguration:
        """Get export configuration for accounts."""
        return ExportConfiguration(
            format='excel',
            include_headers=True,
            include_metadata=True,
            field_mapping={
                'name': 'Account Name',
                'industry': 'Industry',
                'type': 'Account Type',
                'website': 'Website',
                'phone': 'Phone',
                'employees': 'Number of Employees',
                'annual_revenue': 'Annual Revenue',
                'billing_address_street': 'Billing Street',
                'billing_address_city': 'Billing City',
                'billing_address_state': 'Billing State',
                'billing_address_postal_code': 'Billing Postal Code',
                'assigned_to': 'Account Owner'
            }
        )


# Convenience functions
def export_leads(tenant=None, format='excel', filters=None) -> HttpResponse:
    """Export leads with default configuration."""
    from crm.models.lead import Lead
    
    exporter = CRMDataExporter(tenant)
    config = CRMExportTemplates.get_leads_export_config()
    config.format = format
    if filters:
        config.filters = filters
    
    queryset = Lead.objects.filter(tenant=tenant) if tenant else Lead.objects.all()
    return exporter.export_queryset(queryset, config, "leads_export")


def export_opportunities(tenant=None, format='excel', filters=None) -> HttpResponse:
    """Export opportunities with default configuration."""
    from crm.models.opportunity import Opportunity
    
    exporter = CRMDataExporter(tenant)
    config = CRMExportTemplates.get_opportunities_export_config()
    config.format = format
    if filters:
        config.filters = filters
    
    queryset = Opportunity.objects.filter(tenant=tenant) if tenant else Opportunity.objects.all()
    return exporter.export_queryset(queryset, config, "opportunities_export")


def export_contacts(tenant=None, format='excel', filters=None) -> HttpResponse:
    """Export contacts with default configuration."""
    from crm.models.account import Contact
    
    exporter = CRMDataExporter(tenant)
    config = CRMExportTemplates.get_contacts_export_config()
    config.format = format
    if filters:
        config.filters = filters
    
    queryset = Contact.objects.filter(tenant=tenant) if tenant else Contact.objects.all()
    return exporter.export_queryset(queryset, config, "contacts_export")


def export_accounts(tenant=None, format='excel', filters=None) -> HttpResponse:
    """Export accounts with default configuration."""
    from crm.models.account import Account
    
    exporter = CRMDataExporter(tenant)
    config = CRMExportTemplates.get_accounts_export_config()
    config.format = format
    if filters:
        config.filters = filters
    
    queryset = Account.objects.filter(tenant=tenant) if tenant else Account.objects.all()
    return exporter.export_queryset(queryset, config, "accounts_export")


def create_bulk_export(tenant=None, format='excel') -> HttpResponse:
    """Create bulk export of all CRM data."""
    from crm.models.lead import Lead
    from crm.models.opportunity import Opportunity
    from crm.models.account import Account, Contact
    
    exporter = CRMDataExporter(tenant)
    config = ExportConfiguration(format=format, include_metadata=True)
    
    # Prepare data sources
    base_filter = {'tenant': tenant} if tenant else {}
    data_sources = {
        'leads': Lead.objects.filter(**base_filter),
        'opportunities': Opportunity.objects.filter(**base_filter),
        'accounts': Account.objects.filter(**base_filter),
        'contacts': Contact.objects.filter(**base_filter)
    }
    
    return exporter.export_bulk_data(data_sources, config, "crm_bulk_export")